#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Конвертер ВБК XML → XLSX

Преобразует XML файл ВБК (Ведомость банковского контроля) в XLSX
с отдельными листами для каждой таблицы.

Usage:
    python3 scripts/utils/convert_vbk_xml_to_xlsx.py input/Finance/file.xml output/file.xlsx
"""

import sys
import xml.etree.ElementTree as ET
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def parse_vbk_xml(xml_path: Path) -> dict:
    """
    Парсит XML файл ВБК и извлекает данные.
    
    Returns:
        dict с ключами:
            - 'header': словарь с метаинформацией
            - 'resident': данные о резиденте
            - 'tables': список таблиц с данными
    """
    # Автоопределение кодировки: сначала UTF-8, потом windows-1251
    content = None
    for encoding in ['utf-8', 'windows-1251', 'cp1251']:
        try:
            with open(xml_path, 'r', encoding=encoding) as f:
                content = f.read()
            print(f"   📝 Кодировка: {encoding}")
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    
    if content is None:
        raise ValueError(f"Не удалось определить кодировку файла {xml_path}")
    
    root = ET.fromstring(content)
    
    result = {
        'header': {},
        'resident': {},
        'tables': []
    }
    
    # Извлечь заголовок
    theader = root.find('.//THEADER')
    if theader is not None:
        result['header'] = {
            'date': theader.get('date', ''),
            'time': theader.get('time', ''),
            'regn': theader.get('regn', ''),
            'file': theader.get('file', ''),
            'RepType': theader.findtext('RepType', ''),
            'Bank': theader.findtext('Bank', ''),
            'Pasport': theader.findtext('Pasport', ''),
            'Date': theader.findtext('Date', ''),
        }
    
    # Извлечь данные о резиденте
    r1 = root.find('.//R1')
    if r1 is not None:
        result['resident'] = {
            'Resident': r1.findtext('Resident', ''),
            'Subject': r1.findtext('Subject', ''),
            'Rajon': r1.findtext('Rajon', ''),
            'Gorod': r1.findtext('Gorod', ''),
            'NPunkt': r1.findtext('NPunkt', ''),
            'Ulica': r1.findtext('Ulica', ''),
            'Dom': r1.findtext('Dom', ''),
            'Korpus': r1.findtext('Korpus', ''),
            'Ofis': r1.findtext('Ofis', ''),
            'RegNum': r1.findtext('RegNum', ''),
            'RegDate': r1.findtext('RegDate', ''),
            'Inn': r1.findtext('Inn', ''),
        }
    
    # Извлечь все таблицы (Table1, Table2, ...) из всех разделов (R1, R2, ...)
    tbody = root.find('.//TBODY')
    if tbody is not None:
        # Искать во всех Rx разделах
        for rx in tbody:
            if rx.tag.startswith('R'):
                # Найти все элементы, название которых начинается с "Table"
                for child in rx:
                    if child.tag.startswith('Table'):
                        table_name = child.tag
                        n_rec = child.get('nRec', '0')
                        
                        records = []
                        for rec in child.findall('Rec'):
                            rec_data = {'RecID': rec.get('RecID', '')}  # Добавить RecID
                            # Добавить атрибуты записи (date, regn0, etc.)
                            for attr_name, attr_value in rec.attrib.items():
                                if attr_name != 'RecID':  # RecID уже добавлен
                                    rec_data[f'@{attr_name}'] = attr_value
                            # Добавить поля записи
                            for field in rec:
                                rec_data[field.tag] = field.text or ''
                            records.append(rec_data)
                        
                        if records:  # Только если есть записи
                            result['tables'].append({
                                'name': table_name,
                                'nRec': n_rec,
                                'records': records
                            })
    
    return result


def write_xlsx(data: dict, output_path: Path):
    """
    Записывает данные ВБК в XLSX файл.
    """
    wb = Workbook()
    
    # Удалить дефолтный лист
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 1. Лист "Заголовок"
    ws_header = wb.create_sheet("Заголовок")
    ws_header.append(["Параметр", "Значение"])
    
    for key, value in data['header'].items():
        ws_header.append([key, value])
    
    # Форматирование
    for cell in ws_header[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    ws_header.column_dimensions['A'].width = 20
    ws_header.column_dimensions['B'].width = 50
    
    # 2. Лист "Резидент"
    ws_resident = wb.create_sheet("Резидент")
    ws_resident.append(["Параметр", "Значение"])
    
    for key, value in data['resident'].items():
        ws_resident.append([key, value])
    
    # Форматирование
    for cell in ws_resident[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    ws_resident.column_dimensions['A'].width = 20
    ws_resident.column_dimensions['B'].width = 50
    
    # 3. Листы для каждой таблицы
    for table in data['tables']:
        table_name = table['name']
        records = table['records']
        
        if not records:
            continue
        
        # Создать лист
        ws = wb.create_sheet(table_name)
        
        # Заголовки - все уникальные ключи из записей
        headers = []
        for rec in records:
            headers.extend(rec.keys())
        headers = list(dict.fromkeys(headers))  # Уникальные, сохраняя порядок
        
        # Записать заголовки
        ws.append(headers)
        
        # Форматирование заголовков
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Записать данные
        for rec in records:
            row = [rec.get(header, '') for header in headers]
            ws.append(row)
        
        # Автоширина колонок
        for idx, col in enumerate(headers, start=1):
            max_length = len(col)
            for row_idx in range(2, ws.max_row + 1):
                cell_value = str(ws.cell(row_idx, idx).value or '')
                max_length = max(max_length, len(cell_value))
            
            # Ограничить максимальную ширину
            max_length = min(max_length, 50)
            ws.column_dimensions[get_column_letter(idx)].width = max_length + 2
    
    # Сохранить
    wb.save(output_path)
    print(f"✅ XLSX файл создан: {output_path}")
    print(f"   Листов: {len(wb.sheetnames)}")
    print(f"   Листы: {', '.join(wb.sheetnames)}")


def main():
    """
    Основная функция.
    """
    if len(sys.argv) < 2:
        print("Использование: python3 scripts/utils/convert_vbk_xml_to_xlsx.py <input_xml> [output_xlsx]")
        print()
        print("Примеры:")
        print("  python3 scripts/utils/convert_vbk_xml_to_xlsx.py input/Finance/vbk.xml")
        print("  python3 scripts/utils/convert_vbk_xml_to_xlsx.py input/Finance/vbk.xml output/Finance/vbk.xlsx")
        sys.exit(1)
    
    input_path = Path(sys.argv[1])
    
    if not input_path.exists():
        print(f"❌ Файл не найден: {input_path}")
        sys.exit(1)
    
    # Определить output путь
    if len(sys.argv) >= 3:
        output_path = Path(sys.argv[2])
    else:
        # По умолчанию: output/Finance/<имя_файла>.xlsx
        output_dir = Path("output") / "Finance"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / (input_path.stem + ".xlsx")
    
    print(f"📂 Входной файл: {input_path}")
    print(f"📂 Выходной файл: {output_path}")
    print()
    
    # Парсинг XML
    print("📄 Парсинг XML...")
    try:
        data = parse_vbk_xml(input_path)
    except Exception as e:
        print(f"❌ Ошибка парсинга XML: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    
    print(f"   ✅ Заголовок извлечен")
    print(f"   ✅ Резидент извлечен")
    print(f"   ✅ Таблиц найдено: {len(data['tables'])}")
    print()
    
    # Запись в XLSX
    print("📊 Запись в XLSX...")
    try:
        write_xlsx(data, output_path)
    except Exception as e:
        print(f"❌ Ошибка записи XLSX: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    
    print()
    print("🎉 Конвертация завершена успешно!")


if __name__ == "__main__":
    main()

