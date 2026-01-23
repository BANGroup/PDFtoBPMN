#!/usr/bin/env python3
"""
Ручной последовательный парсер PDF НРД
Извлекает данные владельцев СТРОГО В ПОРЯДКЕ ПОЯВЛЕНИЯ в PDF
"""

import re
import fitz
from pathlib import Path
from typing import Optional, List
from dataclasses import dataclass

@dataclass
class OwnerRecord:
    """Запись владельца"""
    page_number: int
    owner_code: str
    full_name: Optional[str] = None
    address: Optional[str] = None
    document_number: Optional[str] = None
    account_number: Optional[str] = None
    quantity: Optional[int] = None

class ManualPDFParser:
    """Последовательный парсер PDF"""
    
    def parse(self, pdf_path: Path) -> List[OwnerRecord]:
        """
        Парсит PDF последовательно, страница за страницей
        ШАГ 1: Извлекаем владельцев
        ШАГ 2: Извлекаем все количества
        ШАГ 3: Сопоставляем по страницам
        """
        print("🔍 ДВУХПРОХОДНЫЙ ПАРСИНГ PDF")
        print("="*80)
        print()
        
        doc = fitz.open(pdf_path)
        
        # ШАГ 1: Извлекаем владельцев и количества в ОДНОМ проходе
        print("📋 ШАГ 1: Извлечение владельцев и количеств...")
        records = []
        used_quantities = set()  # Отслеживаем использованные количества
        
        for page_num in range(len(doc)):
            page = doc[page_num]
            text = page.get_text()
            lines = text.split('\n')
            
            # Ищем владельцев на странице
            owner_matches = list(re.finditer(r'Код\s+(01_\d{10})\(NADC\)', text, re.IGNORECASE))
            
            for match_idx, match in enumerate(owner_matches):
                owner_code = match.group(1)
                record = OwnerRecord(
                    page_number=page_num + 1,
                    owner_code=owner_code
                )
                
                # Извлекаем чанк ПОСЛЕ маркера
                start = match.start()
                next_owner = None
                for next_match in owner_matches:
                    if next_match.start() > start:
                        next_owner = next_match
                        break
                
                chunk_forward = text[start:next_owner.start()] if next_owner else text[start:]
                
                # Извлекаем чанк ДО маркера (для записей где ФИО выше)
                if match_idx > 0:
                    prev_match = owner_matches[match_idx - 1]
                    chunk_back = text[prev_match.start():start]
                else:
                    chunk_back = text[:start]
                
                # Извлекаем поля (сначала пробуем forward, потом back)
                record.full_name = self._extract_fio(chunk_forward) or self._extract_fio(chunk_back)
                record.address = self._extract_address(chunk_forward) or self._extract_address(chunk_back)
                record.document_number = self._extract_document(chunk_forward) or self._extract_document(chunk_back)
                record.account_number = self._extract_account(chunk_forward) or self._extract_account(chunk_back)
                
                # Извлекаем количество - ищем ближайшее к коду владельца
                qty = self._extract_quantity_near_code(lines, match.start(), text, used_quantities, page_num + 1)
                if qty:
                    record.quantity = qty
                    used_quantities.add((page_num + 1, qty))  # Помечаем как использованное
                
                records.append(record)
        
        doc.close()
        
        print(f"   ✅ Найдено владельцев: {len(records)}")
        filled_qty = sum(1 for r in records if r.quantity)
        print(f"   ✅ Заполнено количеств: {filled_qty}/{len(records)}")
        print()
        
        return records
    
    def _extract_quantity_near_code(self, lines: List[str], code_pos: int, full_text: str, used_quantities: set, page_num: int) -> Optional[int]:
        """
        Извлекает количество ближайшее к коду владельца
        Ищет в тексте вокруг позиции code_pos
        """
        # Находим строку с кодом
        text_before_code = full_text[:code_pos]
        code_line_num = text_before_code.count('\n')
        
        # Ищем ближайшую таблицу с рег.номером (ВПЕРЕД и НАЗАД)
        candidates = []
        
        # Поиск вперед (в пределах 50 строк)
        for i in range(code_line_num, min(len(lines), code_line_num + 50)):
            if '4-01-36484-R' in lines[i]:
                # Количество на 2 строки ниже
                if i + 2 < len(lines):
                    qty_line = lines[i + 2].strip()
                    if 'обременен' not in qty_line.lower():
                        qty_match = re.search(r'^\d{1,7}$', qty_line)
                        if qty_match:
                            qty = int(qty_match.group())
                            if 1 <= qty < 3000000:
                                distance = i - code_line_num
                                candidates.append((distance, qty, 'forward'))
        
        # Поиск назад (в пределах 50 строк)
        for i in range(max(0, code_line_num - 50), code_line_num):
            if '4-01-36484-R' in lines[i]:
                # Количество на 2 строки ниже
                if i + 2 < len(lines):
                    qty_line = lines[i + 2].strip()
                    if 'обременен' not in qty_line.lower():
                        qty_match = re.search(r'^\d{1,7}$', qty_line)
                        if qty_match:
                            qty = int(qty_match.group())
                            if 1 <= qty < 3000000:
                                distance = code_line_num - i
                                candidates.append((distance, qty, 'backward'))
        
        # Сортируем по расстоянию (ближайшее первым)
        candidates.sort(key=lambda x: x[0])
        
        # Берем первое НЕиспользованное
        for distance, qty, direction in candidates:
            if (page_num, qty) not in used_quantities:
                return qty
        
        # Если все использованы - берем ближайшее
        if candidates:
            return candidates[0][1]
        
        return None
    
    def _extract_fio(self, chunk: str) -> Optional[str]:
        """Извлекает ФИО из чанка"""
        lines = chunk.split('\n')
        
        # Паттерн 1: "Полное наименование/" (основной)
        for i, line in enumerate(lines):
            if 'Полное наименование' in line:
                # ФИО на 1-2 строки ВЫШЕ
                for offset in [1, 2]:
                    if i - offset >= 0:
                        fio_line = lines[i - offset].strip()
                        
                        # Фильтруем:
                        if fio_line and len(fio_line) > 5:
                            if '(LEI)' not in fio_line and \
                               'Ф.И.О.' not in fio_line and \
                               fio_line not in ['Код', 'Счет', 'Номер'] and \
                               not fio_line.startswith('Номер '):
                                return fio_line
                break
        
        # Паттерн 2: "Ф.И.О. руководителя" (для записей где ФИО выше кода)
        for i, line in enumerate(lines):
            if 'Ф.И.О. руководителя' in line:
                # Название организации на 1-2 строки ВЫШЕ
                for offset in [1, 2]:
                    if i - offset >= 0:
                        fio_line = lines[i - offset].strip()
                        
                        # Фильтруем служебные строки
                        if fio_line and len(fio_line) > 10:
                            if not any(kw in fio_line for kw in ['ОГРН', 'Код ОКПО', 'Код ОКВЭД', 
                                                                  'Регистр', 'Адрес', 'Контакт']):
                                return fio_line
                break
        
        # Паттерн 3: "Адрес для направления корреспонденции" (альтернатива)
        for i, line in enumerate(lines):
            if 'Адрес для направления' in line:
                # Название организации на 2-4 строки ВЫШЕ
                for offset in [2, 3, 4]:
                    if i - offset >= 0:
                        fio_line = lines[i - offset].strip()
                        
                        if fio_line and len(fio_line) > 10:
                            if not any(kw in fio_line for kw in ['ОГРН', 'Код', 'Ф.И.О.', 
                                                                  'Регистр', 'Контакт']):
                                return fio_line
                break
        
        return None
    
    def _extract_address(self, chunk: str) -> Optional[str]:
        """Извлекает адрес"""
        # Паттерн 1: "RU РОССИЯ [индекс] [адрес]" (основной)
        pattern = r'RU\s+РОССИЯ\s+\d{6}\s+([^\n]+?)(?:\s+Адрес|\n)'
        match = re.search(pattern, chunk, re.IGNORECASE)
        if match:
            addr = match.group(1).strip()
            addr = re.sub(r'\s{2,}', ' ', addr)
            if len(addr) > 10:
                return addr
        
        # Паттерн 2: "Адрес для направления корреспонденции" (для chunk_back)
        lines = chunk.split('\n')
        for i, line in enumerate(lines):
            if 'Адрес для направления' in line:
                # Адрес на следующей строке
                if i + 2 < len(lines):
                    addr_line = lines[i + 2].strip()
                    if addr_line and len(addr_line) > 10:
                        # Проверяем что это адрес (есть цифры и запятые)
                        if any(c.isdigit() for c in addr_line):
                            return addr_line
                break
        
        return None
    
    def _extract_document(self, chunk: str) -> Optional[str]:
        """Извлекает номер документа"""
        # Паттерн для физ.лиц: паспорт серия-номер
        pattern1 = r'(?:паспорт|Паспорт).*?(\d{2}\s*\d{2}\s*\d{6})'
        match = re.search(pattern1, chunk, re.IGNORECASE)
        if match:
            doc_num = match.group(1).replace(' ', '')
            # Форматируем: XX XX XXXXXX
            if len(doc_num) == 10:
                return f"{doc_num[:2]} {doc_num[2:4]} {doc_num[4:]}"
        
        # Паттерн для юр.лиц: ОГРН
        pattern2 = r'ОГРН\s+(\d{13})'
        match = re.search(pattern2, chunk, re.IGNORECASE)
        if match:
            return match.group(1)
        
        # Паттерн: ИНН
        pattern3 = r'ИНН\s+(\d{10,12})'
        match = re.search(pattern3, chunk, re.IGNORECASE)
        if match:
            return match.group(1)
        
        return None
    
    def _extract_account(self, chunk: str) -> Optional[str]:
        """Извлекает номер счета"""
        # Паттерн: "Номер [ID] Тип счета"
        pattern = r'Номер\s+([A-Z0-9_/]+)\s+Тип счета'
        match = re.search(pattern, chunk, re.IGNORECASE)
        if match:
            return match.group(1).strip()
        
        return None
    
    def _extract_quantity(self, chunk: str) -> Optional[int]:
        """Извлекает количество облигаций"""
        # Ищем строку с регистрационным номером
        lines = chunk.split('\n')
        
        for i, line in enumerate(lines):
            # Если нашли строку с рег.номером
            if '4-01-36484-R' in line:
                # Пропускаем строки с "обременен"
                if 'обременен' in line.lower():
                    continue
                
                # Смотрим следующие 10 строк
                for j in range(i + 1, min(i + 11, len(lines))):
                    next_line = lines[j].strip()
                    
                    # Ищем число от 1 до 3000000
                    if re.match(r'^\d{1,7}$', next_line):
                        qty = int(next_line)
                        if 1 <= qty < 3000000:
                            return qty
        
        return None


def main():
    """Главная функция"""
    from openpyxl import Workbook
    
    pdf_path = Path("input/Finance/Выпуск 4-01 на 29.07.19.pdf")
    output_path = Path("output/finance/Выпуск_4-01_29_07_19_MANUAL.xlsx")
    
    parser = ManualPDFParser()
    records = parser.parse(pdf_path)
    
    print()
    print("="*80)
    print("📊 СТАТИСТИКА:")
    print("="*80)
    
    total_qty = sum(r.quantity for r in records if r.quantity)
    filled_fio = sum(1 for r in records if r.full_name)
    filled_addr = sum(1 for r in records if r.address)
    filled_doc = sum(1 for r in records if r.document_number)
    
    print(f"   Записей:       {len(records)}")
    print(f"   Облигаций:     {total_qty:,}")
    print(f"   Ожидается:     4,121,600")
    print(f"   Разница:       {total_qty - 4121600:+,}")
    print()
    print("📋 ЗАПОЛНЕННОСТЬ:")
    print(f"   Количество:    {sum(1 for r in records if r.quantity)}/{len(records)}")
    print(f"   ФИО:           {filled_fio}/{len(records)}")
    print(f"   Адрес:         {filled_addr}/{len(records)}")
    print(f"   Документ:      {filled_doc}/{len(records)}")
    print()
    
    # Экспорт в Excel
    print("📦 Создание Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Реестр владельцев"
    
    # Заголовки
    headers = ['Адрес регистрации', 'Количество в штуках', 'Код владельца', 
               'ФИО', 'Номер документа', 'Номер счета', 'Страница']
    ws.append(headers)
    
    # Данные
    for rec in records:
        ws.append([
            rec.address or '',
            rec.quantity or 0,
            rec.owner_code or '',
            rec.full_name or '',
            rec.document_number or '',
            rec.account_number or '',
            rec.page_number or 0
        ])
    
    # Форматирование
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '0'
    
    wb.save(output_path)
    
    print(f"✅ ГОТОВО: {output_path}")
    print()
    print("🎯 Файл создан на основе ПОСЛЕДОВАТЕЛЬНОГО прохода по PDF")
    print("   Порядок записей СООТВЕТСТВУЕТ порядку в PDF")


if __name__ == '__main__':
    main()

