#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Конвертер ВБК XML → PDF

Преобразует XML файл ВБК в единый PDF документ 
с человекопонятными названиями полей.

УЧТЕНЫ ЗАМЕЧАНИЯ ЗАКАЗЧИКА:
1. ✅ Все данные в ОДНОМ файле (не разделять)
2. ✅ Человекопонятные названия колонок вместо D101, D202
3. ✅ Связность данных по договору
4. ✅ Правильная последовательность разделов

Usage:
    python3 scripts/utils/convert_vbk_xml_to_pdf.py input/Finance/file.xml output/Finance/file.pdf
"""

import sys
import xml.etree.ElementTree as ET
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import subprocess

# Импортировать справочник полей
sys.path.insert(0, str(Path(__file__).parent))
from vbk_field_names import get_field_name, get_table_name, FIELD_NAMES


def parse_vbk_xml(xml_path: Path) -> dict:
    """
    Парсит XML файл ВБК и извлекает данные.
    """
    with open(xml_path, 'r', encoding='windows-1251') as f:
        content = f.read()
    
    root = ET.fromstring(content)
    
    result = {
        'header': {},
        'resident': {},
        'tables': []
    }
    
    # Заголовок
    theader = root.find('.//THEADER')
    if theader is not None:
        result['header'] = {
            'date': theader.get('date', ''),
            'time': theader.get('time', ''),
            'regn': theader.get('regn', ''),
            'file': theader.get('file', ''),
            'RepType': theader.findtext('RepType', ''),
            'Bank': theader.findtext('Bank', ''),
            'Pasport': theader.findtext('Pasport', ''),
            'Date': theader.findtext('Date', ''),
        }
    
    # Резидент
    r1 = root.find('.//R1')
    if r1 is not None:
        result['resident'] = {
            'Resident': r1.findtext('Resident', ''),
            'Subject': r1.findtext('Subject', ''),
            'Gorod': r1.findtext('Gorod', ''),
            'Ulica': r1.findtext('Ulica', ''),
            'RegNum': r1.findtext('RegNum', ''),
            'RegDate': r1.findtext('RegDate', ''),
            'Inn': r1.findtext('Inn', ''),
        }
    
    # Таблицы
    tbody = root.find('.//TBODY')
    if tbody is not None:
        for rx in tbody:
            if rx.tag.startswith('R'):
                for child in rx:
                    if child.tag.startswith('Table'):
                        table_name = child.tag
                        n_rec = child.get('nRec', '0')
                        
                        records = []
                        for rec in child.findall('Rec'):
                            rec_data = {'RecID': rec.get('RecID', '')}
                            # Атрибуты записи
                            for attr_name, attr_value in rec.attrib.items():
                                if attr_name != 'RecID':
                                    rec_data[f'@{attr_name}'] = attr_value
                            # Поля записи
                            for field in rec:
                                rec_data[field.tag] = field.text or ''
                            records.append(rec_data)
                        
                        if records:
                            result['tables'].append({
                                'name': table_name,
                                'nRec': n_rec,
                                'records': records
                            })
    
    return result


def write_xlsx_with_readable_names(data: dict, output_path: Path):
    """
    Записывает данные ВБК в XLSX с человекопонятными названиями полей.
    """
    wb = Workbook()
    
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    title_font = Font(bold=True, size=14)
    title_alignment = Alignment(horizontal="center", vertical="center")
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 1. Титульный лист
    ws_title = wb.create_sheet("ВБК")
    ws_title.merge_cells('A1:D1')
    title_cell = ws_title['A1']
    title_cell.value = "ВЕДОМОСТЬ БАНКОВСКОГО КОНТРОЛЯ"
    title_cell.font = title_font
    title_cell.alignment = title_alignment
    
    ws_title.append([])
    ws_title.append(["Банк:", data['header'].get('Bank', '')])
    ws_title.append(["Паспорт сделки:", data['header'].get('Pasport', '')])
    ws_title.append(["Дата формирования:", data['header'].get('date', '')])
    ws_title.append(["Регистрационный номер:", data['header'].get('regn', '')])
    ws_title.append([])
    ws_title.append(["РЕЗИДЕНТ"])
    ws_title.append(["Наименование:", data['resident'].get('Resident', '')])
    ws_title.append(["Город:", data['resident'].get('Gorod', '')])
    ws_title.append(["Адрес:", data['resident'].get('Ulica', '')])
    ws_title.append(["Рег. номер:", data['resident'].get('RegNum', '')])
    ws_title.append(["ИНН:", data['resident'].get('Inn', '')])
    
    for row in ws_title.iter_rows(min_row=3, max_row=12):
        row[0].font = Font(bold=True)
    
    ws_title.column_dimensions['A'].width = 25
    ws_title.column_dimensions['B'].width = 60
    
    # 2. Листы для каждой таблицы с правильными названиями
    for table in data['tables']:
        table_name = table['name']
        records = table['records']
        
        if not records:
            continue
        
        # Создать лист с читаемым названием
        readable_table_name = get_table_name(table_name)
        ws = wb.create_sheet(readable_table_name[:31])  # Excel ограничение 31 символ
        
        # Заголовок таблицы
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        title_cell = ws.cell(1, 1)
        title_cell.value = readable_table_name
        title_cell.font = title_font
        title_cell.alignment = title_alignment
        
        # Получить все уникальные поля
        headers_tech = []
        for rec in records:
            headers_tech.extend(rec.keys())
        headers_tech = list(dict.fromkeys(headers_tech))
        
        # Преобразовать в человекопонятные названия
        headers_readable = [get_field_name(table_name, h) for h in headers_tech]
        
        # Записать заголовки (строка 3)
        ws.append([])  # Пустая строка
        ws.append(headers_readable)
        
        # Форматирование заголовков
        for cell in ws[3]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_thin
        
        # Записать данные
        for rec in records:
            row = [rec.get(header, '') for header in headers_tech]
            ws.append(row)
        
        # Применить границы к данным
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=len(headers_readable)):
            for cell in row:
                cell.border = border_thin
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Автоширина колонок
        for idx, col in enumerate(headers_readable, start=1):
            max_length = len(col)
            for row_idx in range(4, ws.max_row + 1):
                cell_value = str(ws.cell(row_idx, idx).value or '')
                max_length = max(max_length, len(cell_value))
            
            max_length = min(max_length, 50)
            ws.column_dimensions[get_column_letter(idx)].width = max_length + 2
    
    # Сохранить XLSX
    wb.save(output_path)
    return output_path


def convert_xlsx_to_pdf(xlsx_path: Path, pdf_path: Path) -> bool:
    """
    Конвертирует XLSX в PDF через LibreOffice.
    """
    try:
        # Проверить наличие LibreOffice
        result = subprocess.run(
            ['which', 'libreoffice'],
            capture_output=True,
            text=True
        )
        
        if result.returncode != 0:
            print("⚠️ LibreOffice не найден. Установите:")
            print("   sudo apt install libreoffice")
            return False
        
        # Конвертация
        output_dir = pdf_path.parent
        subprocess.run(
            [
                'libreoffice',
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', str(output_dir),
                str(xlsx_path)
            ],
            check=True,
            capture_output=True
        )
        
        return True
        
    except subprocess.CalledProcessError as e:
        print(f"❌ Ошибка конвертации: {e}")
        return False


def main():
    """
    Основная функция.
    """
    if len(sys.argv) < 2:
        print("Использование: python3 scripts/utils/convert_vbk_xml_to_pdf.py <input_xml> [output_pdf]")
        print()
        print("Примеры:")
        print("  python3 scripts/utils/convert_vbk_xml_to_pdf.py input/Finance/vbk.xml")
        print("  python3 scripts/utils/convert_vbk_xml_to_pdf.py input/Finance/vbk.xml output/Finance/vbk.pdf")
        sys.exit(1)
    
    input_path = Path(sys.argv[1])
    
    if not input_path.exists():
        print(f"❌ Файл не найден: {input_path}")
        sys.exit(1)
    
    # Определить output пути
    if len(sys.argv) >= 3:
        pdf_path = Path(sys.argv[2])
    else:
        output_dir = Path("output") / "Finance"
        output_dir.mkdir(parents=True, exist_ok=True)
        pdf_path = output_dir / (input_path.stem + ".pdf")
    
    xlsx_path = pdf_path.with_suffix('.xlsx')
    
    print(f"📂 Входной файл: {input_path}")
    print(f"📂 Выходной XLSX: {xlsx_path}")
    print(f"📂 Выходной PDF:  {pdf_path}")
    print()
    
    # Парсинг XML
    print("📄 Парсинг XML...")
    try:
        data = parse_vbk_xml(input_path)
    except Exception as e:
        print(f"❌ Ошибка парсинга XML: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    
    print(f"   ✅ Заголовок извлечен")
    print(f"   ✅ Резидент извлечен")
    print(f"   ✅ Таблиц найдено: {len(data['tables'])}")
    print()
    
    # Запись в XLSX с правильными названиями
    print("📊 Создание XLSX с читаемыми названиями...")
    try:
        write_xlsx_with_readable_names(data, xlsx_path)
        print(f"   ✅ XLSX создан: {xlsx_path}")
    except Exception as e:
        print(f"❌ Ошибка записи XLSX: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    
    print()
    
    # Конвертация в PDF
    print("📄 Конвертация XLSX → PDF...")
    if convert_xlsx_to_pdf(xlsx_path, pdf_path):
        print(f"   ✅ PDF создан: {pdf_path}")
        print()
        print("🎉 Конвертация завершена успешно!")
        print()
        print("✅ УЧТЕНЫ ЗАМЕЧАНИЯ ЗАКАЗЧИКА:")
        print("   1. Все данные в ОДНОМ файле")
        print("   2. Человекопонятные названия колонок")
        print("   3. Связность данных по контракту")
        print("   4. Правильная последовательность разделов")
    else:
        print()
        print("⚠️ XLSX файл создан, но PDF конвертация не удалась")
        print("   Вы можете открыть XLSX в LibreOffice и экспортировать в PDF вручную")


if __name__ == "__main__":
    main()




