#!/usr/bin/env python3
"""
Гибридный парсер для VBK документа.
Использует find_tables() для структуры + группировку строк по номерам п/п.
Интегрирован с NameNormalizer для устранения дубликатов контрагентов.
"""

import fitz
import pandas as pd
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import numbers
from typing import List

# Импорт нормализатора
from name_normalizer import get_normalizer


class VBKHybridParser:
    """Гибридный парсер: структура из find_tables + группировка по номерам"""
    
    def __init__(self, pdf_path: str, section: str = "II"):
        self.pdf_path = pdf_path
        self.section = section
        self.doc = None
        
        if section == "II":
            self.num_columns = 19
            self.financial_col_indices = [6, 8]  # 0-indexed (колонки с суммами)
            self.headers = [
                "№ п/п",                                          # 1
                "Дата операции",                                   # 2
                "Направление (признак) платежа",                  # 3
                "Подколонка 3a",                                  # 3a
                "Код вида операции",                              # 4
                "Код валюты (платеж)",                            # 5
                "Сумма (платеж)",                                 # 6
                "Код валюты (контракт)",                          # 7
                "Сумма (контракт)",                               # 8
                "Колонка 9",                                      # 9
                "Ожидаемый срок репатриации",                     # 10
                "Колонка 11",                                     # 11
                "Колонка 12",                                     # 12
                "Колонка 12a",                                    # 12a
                "Колонка 12б",                                    # 12б
                "Колонка 12в",                                    # 12в
                "Колонка 13",                                     # 13
                "Текстовое поле (описание операции)",             # 14
                "Дополнительная информация"                        # 15
            ]
        else:  # section == "III"
            self.num_columns = 15
            self.financial_col_indices = [6, 8]
            self.headers = [
                "№ п/п",
                "Подтверждающий документ - номер",
                "Подтверждающий документ - дата",
                "Код вида подтверждающего документа",
                "Признак исполнения обязательств третьим лицом",
                "Сумма по документам (документ) - код валюты",
                "Сумма по документам (документ) - сумма",
                "Сумма по документам (контракт) - код валюты",
                "Сумма по документам (контракт) - сумма",
                "Признак поставки",
                "Срок исполнения",
                "Признак изменения записи",
                "Код страны грузоотправителя/грузополучателя",
                "Дополнительная информация",
                "Примечание"
            ]
    
    def parse(self):
        """Извлечь данные из PDF"""
        self.doc = fitz.open(self.pdf_path)
        
        print(f"📄 Обработка файла: {self.pdf_path}")
        print(f"📊 Раздел: {self.section}")
        print(f"📊 Всего страниц: {len(self.doc)}")
        print()
        
        # Находим стартовую страницу
        start_page = self._find_section_start()
        if start_page is None:
            print(f"❌ Не найден 'Раздел {self.section}' в документе")
            return pd.DataFrame()
        
        print(f"✅ Раздел {self.section} найден на странице {start_page + 1}")
        print()
        
        # ЭТАП 1: Извлекаем все строки через find_tables()
        all_table_rows = []
        page_num = start_page
        
        while page_num < len(self.doc):
            page = self.doc[page_num]
            
            tab_finder = page.find_tables()
            tables = tab_finder.tables
            
            if not tables:
                page_num += 1
                continue
            
            table = tables[0]
            df = table.to_pandas()
            
            # Проверяем структуру таблицы
            num_cols = len(df.columns)
            expected_cols = 19 if self.section == "II" else 15
            other_cols = 15 if self.section == "II" else 19
            
            if abs(num_cols - other_cols) < abs(num_cols - expected_cols):
                # Структура изменилась - конец раздела
                print(f"  ⚠️ Конец Раздела {self.section} на странице {page_num + 1}")
                break
            
            # Переименовываем колонки в числовые индексы
            df.columns = range(len(df.columns))
            
            # Добавляем все строки (включая без номера п/п)
            for idx, row in df.iterrows():
                first_col = str(row.iloc[0]).strip()
                
                # Пропускаем пустые и заголовки
                if not first_col or first_col in ['None', '', 'nan', '№ п/п']:
                    continue
                
                # СНАЧАЛА исправляем смещение колонок (если есть)
                # Раздел II: Col 1 пустая, Col 2 = дата → удаляем Col 1
                # Раздел III: Col 1 ВСЕГДА пустая (артефакт) → удаляем Col 1
                if len(row) > 2:
                    col0_val = str(row.iloc[0]).strip()
                    col1_val = str(row.iloc[1]).strip()
                    col2_val = str(row.iloc[2]).strip()
                    
                    is_row_number = re.match(r'^\d{1,4}$', col0_val)
                    
                    # Для Раздела III: всегда удалять Col 1 если пустая
                    if self.section == "III" and is_row_number and col1_val in ['None', 'nan', '']:
                        # Удаляем Col 1: [Col0, Col1, Col2, ...] → [Col0, Col2, ...]
                        row = pd.concat([row.iloc[:1], row.iloc[2:]], ignore_index=True)
                        # Дополняем до нужной длины
                        row = pd.concat([row, pd.Series([None])], ignore_index=True)
                    
                    # Для Раздела II: удалять Col 1 только если Col 2 = дата
                    elif self.section == "II" and is_row_number and col1_val in ['None', 'nan', ''] and re.match(r'\d{2}\.\d{2}\.\d{4}', col2_val):
                        # Удаляем Col 1: [Col0, Col1, Col2, ...] → [Col0, Col2, ...]
                        row = pd.concat([row.iloc[:1], row.iloc[2:]], ignore_index=True)
                        # Дополняем до нужной длины
                        row = pd.concat([row, pd.Series([None])], ignore_index=True)
                    
                    # Случай для строк без номера (продолжение записи)
                    elif not is_row_number and col1_val in ['None', 'nan', ''] and re.match(r'\d{2}\.\d{2}\.\d{4}', col2_val):
                        row = pd.concat([row.iloc[:1], row.iloc[2:]], ignore_index=True)
                        row = pd.concat([row, pd.Series([None])], ignore_index=True)
                
                # ПОТОМ проверяем фильтры (на исправленной строке)
                # Пропускаем номера колонок (только если нет даты в нужной колонке ПОСЛЕ исправления)
                first_col_after_fix = str(row.iloc[0]).strip()
                if first_col_after_fix in ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', 
                                             '11', '12', '13', '14', '15', '16', '17', '18', '19'] and idx < 5:
                    # После удаления артефакт-колонки:
                    # Раздел II:  [№, Дата, ...]  → дата в Col 1
                    # Раздел III: [№, Номер док, Дата док, ...]  → дата в Col 2
                    if self.section == "II":
                        date_col = str(row.iloc[1] if len(row) > 1 else '').strip()
                    else:  # section == "III"
                        date_col = str(row.iloc[2] if len(row) > 2 else '').strip()
                    
                    # Если дата не найдена - это заголовок, пропускаем
                    if not re.match(r'\d{2}\.\d{2}\.\d{4}', date_col):
                        continue
                
                # Обрезаем/дополняем до нужного количества колонок
                if len(row) > self.num_columns:
                    row = row[:self.num_columns]
                elif len(row) < self.num_columns:
                    for i in range(len(row), self.num_columns):
                        row[i] = None
                
                all_table_rows.append((page_num + 1, row))
            
            page_num += 1
        
        print(f"📍 Раздел {self.section}: страницы {start_page + 1} - {page_num}")
        print(f"📝 Извлечено строк таблицы: {len(all_table_rows)}")
        print()
        
        self.doc.close()
        
        # ЭТАП 2: Группируем строки по номерам п/п
        grouped_records = self._group_rows_by_numbers(all_table_rows)
        
        # ЭТАП 3: Создаем DataFrame
        df = self._build_dataframe(grouped_records)
        
        print(f"\n📊 Извлечено записей: {len(df)}")
        
        return df
    
    def _find_section_start(self) -> int:
        """Найти страницу с началом раздела"""
        marker = f"Раздел {self.section}"
        for page_num in range(len(self.doc)):
            page = self.doc[page_num]
            text = page.get_text()
            if marker in text:
                if self.section == "III":
                    return page_num + 1 if page_num + 1 < len(self.doc) else page_num
                return page_num
        return None
    
    def _group_rows_by_numbers(self, all_rows: List[tuple]) -> dict:
        """
        Группировать строки по номерам п/п.
        
        Алгоритм:
        1. Если строка начинается с числа - это новая запись
        2. Если строка НЕ начинается с числа - это продолжение предыдущей
        3. Склеиваем значения колонок (особенно текстовые поля)
        """
        grouped = {}
        current_row_num = None
        current_rows = []
        
        print("🔄 Группировка строк по номерам п/п...")
        
        for page_num, row in all_rows:
            first_col = str(row.iloc[0]).strip()
            
            # Проверяем, является ли это номером п/п
            is_row_number = re.match(r'^\d{1,4}$', first_col)
            
            if is_row_number:
                # Это новая запись
                row_num = int(first_col)
                
                # Сохраняем предыдущую запись
                if current_row_num is not None:
                    grouped[current_row_num] = self._merge_rows(current_rows)
                
                # Начинаем новую запись
                current_row_num = row_num
                current_rows = [row]
            else:
                # Это продолжение текущей записи
                if current_row_num is not None:
                    current_rows.append(row)
        
        # Сохраняем последнюю запись
        if current_row_num is not None:
            grouped[current_row_num] = self._merge_rows(current_rows)
        
        print(f"  ✅ Сгруппировано записей: {len(grouped)}")
        
        return grouped
    
    def _merge_rows(self, rows: List[pd.Series]) -> pd.Series:
        """
        Объединить несколько строк в одну.
        Склеивает текстовые значения, сохраняет числовые из первой строки.
        """
        if len(rows) == 1:
            return rows[0]
        
        # Базовая строка - первая (с номером п/п)
        merged = rows[0].copy()
        
        # Для каждой колонки
        for col_idx in range(len(merged)):
            values = []
            
            for row in rows:
                if col_idx < len(row):
                    val = str(row.iloc[col_idx]).strip()
                    if val and val not in ['None', 'nan', '']:
                        values.append(val)
            
            # Если это числовая колонка (сумма, код) - берем первое непустое значение
            if col_idx in [4, 5, 6, 7, 8, 9]:  # Коды и суммы
                merged.iloc[col_idx] = values[0] if values else None
            else:
                # Текстовые колонки - склеиваем через пробел
                if len(values) > 1:
                    merged.iloc[col_idx] = ' '.join(values)
                elif values:
                    merged.iloc[col_idx] = values[0]
        
        return merged
    
    def _build_dataframe(self, grouped: dict) -> pd.DataFrame:
        """Построить DataFrame из сгруппированных записей"""
        rows = []
        
        for row_num in sorted(grouped.keys()):
            row = grouped[row_num]
            
            # Преобразуем в список значений
            row_data = [row.iloc[i] if i < len(row) else None for i in range(self.num_columns)]
            rows.append(row_data)
        
        df = pd.DataFrame(rows, columns=self.headers)
        
        # Преобразуем типы
        df = self._convert_types(df)
        
        return df
    
    def _convert_types(self, df: pd.DataFrame) -> pd.DataFrame:
        """Преобразовать типы данных"""
        # Преобразуем финансовые колонки
        for col_idx in self.financial_col_indices:
            col_name = self.headers[col_idx]
            df[col_name] = df[col_name].apply(self._parse_number)
        
        # Очищаем пустые значения
        df = df.replace(['None', 'nan', ''], None)
        
        return df
    
    def _parse_number(self, value):
        """Преобразовать строку в число"""
        if pd.isna(value) or value is None or value == '':
            return None
        
        value_str = str(value).strip()
        if not value_str or value_str in ['None', 'nan', '']:
            return None
        
        # Убираем пробелы и заменяем запятую на точку
        value_str = value_str.replace(' ', '').replace(',', '')
        
        try:
            return float(value_str)
        except ValueError:
            return None
    
    def _normalize_names(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Добавляет нормализованные имена для контрагентов.
        
        Для Раздела II: "Наименование отправителя (плательщика) / получателя" (индекс 12)
        Для Раздела III: "Наименование плательщика-1" (индекс 8), "Наименование получателя-1" (индекс 9)
        """
        normalizer = get_normalizer()
        
        if self.section == "II":
            # Колонка с наименованием контрагента
            name_col = "Наименование отправителя (плательщика) / получателя"
            if name_col in df.columns:
                # Создаём нормализованную версию
                df[f"{name_col} (норм.)"] = df[name_col].apply(
                    lambda x: normalizer.normalize(str(x)) if pd.notna(x) else ''
                )
        else:  # section == "III"
            # Две колонки с наименованиями
            payer_col = "Наименование плательщика-1"
            receiver_col = "Наименование получателя-1"
            
            if payer_col in df.columns:
                df[f"{payer_col} (норм.)"] = df[payer_col].apply(
                    lambda x: normalizer.normalize(str(x)) if pd.notna(x) else ''
                )
            
            if receiver_col in df.columns:
                df[f"{receiver_col} (норм.)"] = df[receiver_col].apply(
                    lambda x: normalizer.normalize(str(x)) if pd.notna(x) else ''
                )
        
        return df
    
    def save_to_excel(self, df: pd.DataFrame, output_path: str):
        """Сохранить в Excel с правильными форматами и нормализованными именами"""
        # Добавляем нормализованные имена
        df = self._normalize_names(df)
        
        df.to_excel(output_path, index=False, engine='openpyxl')
        
        # Применяем форматирование
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Применяем финансовый формат
        for col_idx in self.financial_col_indices:
            excel_col = col_idx + 1  # Excel 1-indexed
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=excel_col)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = numbers.FORMAT_NUMBER_00
        
        wb.save(output_path)
        print(f"\n💾 Сохранено в: {output_path}")
        print(f"   Применен финансовый формат для {len(self.financial_col_indices)} колонок")
        print(f"   Добавлены нормализованные имена контрагентов")

