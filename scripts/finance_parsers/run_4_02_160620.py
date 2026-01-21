#!/usr/bin/env python3
"""
Скрипт для обработки: Выпуск 4-02 на 16.06.2020
Регистрационный номер: 4-02-36484-R
"""

import sys
import re
from pathlib import Path
from dataclasses import dataclass
from typing import Optional, List
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from scripts.finance_parser.md_parser_4_02 import MDParser
from scripts.finance_parser.models import OwnerRecord

def main():
    """Главная функция"""
    print("="*80)
    print("🔍 ОБРАБОТКА: Выпуск 4-02 на 16.06.2020")
    print("📋 Регистрационный номер: 4-02-36484-R")
    print("="*80)
    print()
    
    # Пути к файлам
    md_path = Path("output/Выпуск_4-02_на_16.06.2020/Выпуск_4-02_на_16.06.2020_OCR.md")
    output_path = Path("output/finance/Выпуск_4-02_на_16.06.2020.xlsx")
    
    # Создаем output директорию
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Парсинг MD
    print("📖 Парсинг MD файла...")
    parser = MDParser()
    records = parser.parse_md_file(str(md_path))
    print(f"   ✅ Извлечено записей: {len(records)}")
    print()
    
    # Валидация
    print("🔍 Валидация данных...")
    total_bonds = sum(r.quantity for r in records if r.quantity)
    filled_qty = sum(1 for r in records if r.quantity)
    filled_fio = sum(1 for r in records if r.full_name)
    filled_addr = sum(1 for r in records if r.address)
    filled_doc = sum(1 for r in records if r.document_number)
    
    print(f"   Записей: {len(records)}")
    print(f"   Облигаций: {total_bonds:,}")
    print()
    print("   Заполненность:")
    print(f"   • Количество:  {100*filled_qty/len(records):.1f}%")
    print(f"   • ФИО:         {100*filled_fio/len(records):.1f}%")
    print(f"   • Адрес:       {100*filled_addr/len(records):.1f}%")
    print(f"   • Документ:    {100*filled_doc/len(records):.1f}%")
    print()
    
    # Экспорт в Excel
    print("📦 Создание Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Реестр владельцев"
    
    # Заголовки
    headers = ['Адрес регистрации', 'Количество в штуках', 'Код владельца', 
               'ФИО', 'Номер документа', 'Номер счета', 'Страница']
    ws.append(headers)
    
    # Данные
    for rec in records:
        ws.append([
            rec.address or '',
            rec.quantity or 0,
            rec.owner_code or '',
            rec.full_name or '',
            rec.document_number or '',
            rec.account_number or '',
            rec.page_number or 0
        ])
    
    # Форматирование
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '0'
    
    wb.save(output_path)
    
    print(f"✅ ГОТОВО: {output_path}")
    print()
    
    # Итоговая статистика
    print("="*80)
    print("📊 ИТОГОВАЯ СТАТИСТИКА")
    print("="*80)
    print(f"   Записей:       {len(records)}")
    print(f"   Облигаций:     {total_bonds:,}")
    print()
    print(f"   Заполненность:")
    print(f"   • Количество:  {100*filled_qty/len(records):.1f}%")
    print(f"   • ФИО:         {100*filled_fio/len(records):.1f}%")
    print(f"   • Адрес:       {100*filled_addr/len(records):.1f}%")
    print(f"   • Документ:    {100*filled_doc/len(records):.1f}%")
    print()
    print("="*80)


if __name__ == '__main__':
    main()

