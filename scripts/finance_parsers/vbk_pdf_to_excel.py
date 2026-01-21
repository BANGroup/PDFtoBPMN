#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Конвертация ВБК PDF → XLSX

Требования заказчика:
1. Один XLSX файл на весь ВБК
2. Каждый раздел ВБК — отдельный лист книги
3. Человекопонятные заголовки (никаких «Колонка 9»)
4. Структура листов совпадает с эталонным Excel из банка

Usage:
    python3 scripts/utils/vbk_pdf_to_excel.py input/VBK.pdf [output.xlsx]
"""

import argparse
import re
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Чтобы импортировать pdf_to_context
import sys

CURRENT_DIR = Path(__file__).resolve().parent
ROOT_DIR = CURRENT_DIR.parent.parent
sys.path.insert(0, str(ROOT_DIR / "scripts"))

from pdf_to_context.pipeline import PDFToContextPipeline
from pdf_to_context.models.data_models import ContentType

try:
    import xlrd2  # type: ignore
except ImportError:
    xlrd2 = None  # Optional: требуется только для шаблонов заголовков


# ------------------------------------------------------------
# Конфигурация разделов
# ------------------------------------------------------------
SECTION_PATTERNS: List[Tuple[re.Pattern, str]] = [
    (re.compile(r"подраздел\s+III\.II", re.IGNORECASE), "Раздел III. Подраздел III.II"),
    (re.compile(r"подраздел\s+III\.I", re.IGNORECASE), "Раздел III. Подраздел III.I"),
    (re.compile(r"подраздел\s+IV\.II", re.IGNORECASE), "Раздел IV. Подраздел IV.II"),
    (re.compile(r"подраздел\s+IV\.I", re.IGNORECASE), "Раздел IV. Подраздел IV.I"),
    (re.compile(r"раздел\s+I\b", re.IGNORECASE), "Раздел I"),
    (re.compile(r"раздел\s+II\b", re.IGNORECASE), "Раздел II"),
    (re.compile(r"раздел\s+III\b", re.IGNORECASE), "Раздел III. Подраздел III.I"),
    (re.compile(r"раздел\s+IV\b", re.IGNORECASE), "Раздел IV. Подраздел IV.I"),
    (re.compile(r"раздел\s+V\b", re.IGNORECASE), "Раздел V"),
]

SECTION_ORDER = [
    "Раздел I",
    "Раздел II",
    "Раздел III. Подраздел III.I",
    "Раздел III. Подраздел III.II",
    "Раздел IV. Подраздел IV.I",
    "Раздел IV. Подраздел IV.II",
    "Раздел V",
]

TITLE_SHEET = "Титульный лист"
SECTION_HEADER_KEYWORDS: List[Tuple[str, List[str]]] = [
    (
        "Раздел II",
        [
            "дата операции",
            "направление (признак) платежа",
            "код вида операции",
            "сумма операции",
        ],
    ),
    (
        "Раздел III. Подраздел III.I",
        [
            "подтверждающий документ",
            "код вида подтверждающего документа",
            "сумма по подтверждающим документам",
        ],
    ),
    (
        "Раздел III. Подраздел III.II",
        [
            "признак оформления дт",
            "код вида подтверждающего документа",
        ],
    ),
    (
        "Раздел IV. Подраздел IV.I",
        [
            "подтверждающий документ",
            "код вида подтверждающего документа",
            "признак поставки",
        ],
    ),
    (
        "Раздел V",
        [
            "сальдо расчетов",
            "итоговые данные расчетов по контракту",
        ],
    ),
]

COMBINED_SECTIONS = {
    "Раздел II",
    "Раздел III. Подраздел III.I",
    "Раздел III. Подраздел III.II",
    "Раздел IV. Подраздел IV.I",
    "Раздел IV. Подраздел IV.II",
    "Раздел V",
}
NUMBER_FORMAT = "# ##0,00"
NUMBER_PATTERN = re.compile(
    r"^-?\d{1,3}(?:[ ,\u00A0]\d{3})*(?:[.,]\d+)?$|^-?\d+(?:[.,]\d+)?$"
)


def normalize_section(text: str) -> str:
    """Определить название листа по заголовку."""
    cleaned = text.replace("\n", " ").strip()
    for pattern, name in SECTION_PATTERNS:
        if pattern.search(cleaned):
            return name

    # Заголовки типа "7. Особые условия..." относим к Разделу I
    if re.search(r"^\d+\.", cleaned):
        return "Раздел I"

    return ""


def extract_tables_with_sections(pdf_path: Path) -> Tuple[List[Dict], Dict[str, List[Dict]]]:
    """Извлечь таблицы и распределить по разделам."""
    pipeline = PDFToContextPipeline(
        enable_ocr=False,
        extract_images=False,
        extract_drawings=False,
        extract_tables=True,
        include_frontmatter=False,
        include_toc=False,
    )
    ir = pipeline.process_to_ir(str(pdf_path))
    reading_order = ir.get_reading_order()

    section_by_table_id: Dict[str, str] = {}
    current_section = "Раздел I"

    for block in reading_order:
        if block.type in (ContentType.HEADING, ContentType.PARAGRAPH):
            section_name = normalize_section(block.content)
            if section_name:
                current_section = section_name
        if block.type == ContentType.TABLE:
            section_by_table_id[block.id] = current_section

    tables = [b for b in ir.blocks if b.type == ContentType.TABLE]

    grouped: Dict[str, List[Dict]] = {}
    for table in tables:
        section = detect_section_by_header(table.metadata.get("data") or []) or section_by_table_id.get(table.id, "Раздел I")
        data = table.metadata.get("data") or []
        if not data:
            continue
        clean_rows = [[_normalize_cell(cell) for cell in row] for row in data]
        grouped.setdefault(section, []).append(
            {
                "page": table.page,
                "data": clean_rows,
            }
        )
    return tables, grouped


def detect_section_by_header(data: List[List[str]]) -> str:
    for row in data:
        normalized = " ".join(_normalize_cell(cell).lower() for cell in row if _normalize_cell(cell))
        if not normalized:
            continue
        for section, keywords in SECTION_HEADER_KEYWORDS:
            if all(keyword in normalized for keyword in keywords):
                return section
    return ""


def load_header_templates() -> Dict[str, List[List[str]]]:
    """Подгрузить образец заголовков из эталонного XLS, если доступен."""
    templates: Dict[str, List[List[str]]] = {}
    sample_path = Path("input/Finance/16060002_9.xls")
    if not sample_path.exists() or xlrd2 is None:
        return templates

    book = xlrd2.open_workbook(str(sample_path))
    for sheet_name in book.sheet_names():
        sheet = book.sheet_by_name(sheet_name)
        rows: List[List[str]] = []
        numbering_seen = False
        for r in range(min(sheet.nrows, 40)):
            row = [_normalize_cell(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
            if any(cell for cell in row):
                rows.append(row)
                if any(re.fullmatch(r"\d+(\s*\([^)]+\))?", cell) for cell in row if cell):
                    numbering_seen = True
                    break
            elif rows:
                break
        if rows and numbering_seen:
            templates[sheet_name] = rows
    return templates


def is_header_like(row: List[str]) -> bool:
    if not any(row):
        return True
    row_join = " ".join(cell.lower() for cell in row if cell)
    if "№" in row_join or "код" in row_join and not any(ch.isdigit() for ch in row_join):
        return True
    digits = [cell for cell in row if re.fullmatch(r"\d+(\s*\([^)]+\))?", cell or "")]
    return len(digits) >= 5 and digits[0] == "1"


def combine_section_rows(
    tables: List[Dict], keep_first_header: bool, max_cols: int
) -> List[List]:
    combined: List[List] = []
    header_saved = False
    for tbl in tables:
        for row in tbl["data"]:
            header = is_header_like(row)
            if header:
                if keep_first_header and not header_saved:
                    combined.append(
                        _prepare_row(row, max_cols, convert_numbers=False)
                    )
                    header_saved = True
                continue
            combined.append(_prepare_row(row, max_cols, convert_numbers=True))
    return combined


def _normalize_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value}"
    return str(value).strip()


def _prepare_row(row: List[str], max_cols: int, convert_numbers: bool) -> List:
    normalized = [_normalize_cell(cell) for cell in row]
    if len(normalized) < max_cols:
        normalized.extend([""] * (max_cols - len(normalized)))
    else:
        normalized = normalized[:max_cols]
    if convert_numbers:
        normalized = [_convert_to_number(cell) for cell in normalized]
    return normalized


def _convert_to_number(value):
    if not isinstance(value, str):
        return value
    stripped = value.strip()
    if not stripped:
        return ""
    cleaned = stripped.replace("\u00A0", "").replace(" ", "")
    if not NUMBER_PATTERN.match(cleaned):
        return value

    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(".") > cleaned.rfind(","):
            cleaned = cleaned.replace(",", "")
        else:
            cleaned = cleaned.replace(".", "").replace(",", ".")
    else:
        cleaned = cleaned.replace(",", ".")

    try:
        return float(cleaned)
    except ValueError:
        return value


def build_workbook(grouped_tables: Dict[str, List[Dict]], output_path: Path):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Титульный лист с краткой справкой
    ws_title = wb.create_sheet(TITLE_SHEET)
    ws_title["A1"] = "ВЕДОМОСТЬ БАНКОВСКОГО КОНТРОЛЯ (автоматическая выгрузка)"
    ws_title["A1"].font = Font(bold=True, size=14)
    ws_title.merge_cells("A1:D1")
    ws_title["A2"] = "Секции:"
    ws_title["A2"].font = Font(bold=True)
    row = 3
    for section in SECTION_ORDER:
        count = len(grouped_tables.get(section, []))
        ws_title[f"A{row}"] = section
        ws_title[f"B{row}"] = f"{count} таблиц"
        row += 1

    header_templates = load_header_templates()

    for section in SECTION_ORDER:
        tables = grouped_tables.get(section)
        if not tables:
            continue
        sheet_name = section[:31]
        ws = wb.create_sheet(sheet_name)
        ws.append([section])
        ws["A1"].font = Font(bold=True, size=13)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

        if section in COMBINED_SECTIONS:
            ws.append([])
            header_rows = header_templates.get(section, [])
            max_cols = max(
                max((len(row) for row in header_rows), default=0),
                max(
                    (len(r) for tbl in tables for r in tbl["data"]),
                    default=0,
                ),
            )
            if max_cols == 0:
                continue

            for row in header_rows:
                prepared = _prepare_row(row, max_cols, convert_numbers=False)
                ws.append(prepared)
                row_idx = ws.max_row
                for col_idx in range(1, max_cols + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border_thin
            combined_rows = combine_section_rows(
                tables, keep_first_header=not header_rows, max_cols=max_cols
            )
            for row_data in combined_rows:
                ws.append(row_data)
                row_idx = ws.max_row
                for col_idx in range(1, max_cols + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = border_thin
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    if isinstance(cell.value, float):
                        cell.number_format = NUMBER_FORMAT
            for idx_col in range(1, max_cols + 1):
                max_len = 0
                for r in range(2, ws.max_row + 1):
                    value = ws.cell(row=r, column=idx_col).value
                    if value:
                        max_len = max(max_len, min(len(str(value)), 50))
                ws.column_dimensions[get_column_letter(idx_col)].width = max(10, max_len + 2)
            continue

        for idx, table in enumerate(tables, start=1):
            ws.append([])
            ws.append([f"Таблица {idx} (стр. {table['page']})"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

            data = table["data"]
            if not data:
                continue
            max_cols = max(len(row) for row in data)
            for row_data in data:
                header_row = is_header_like(row_data)
                prepared = _prepare_row(
                    row_data, max_cols, convert_numbers=not header_row
                )
                ws.append(prepared)
                row_idx = ws.max_row
                for col_idx in range(1, max_cols + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = border_thin
                    if header_row:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    else:
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                        if isinstance(cell.value, float):
                            cell.number_format = NUMBER_FORMAT

            start_row = 3
            for idx_col in range(1, max_cols + 1):
                max_len = 0
                for r in range(start_row, ws.max_row + 1):
                    value = ws.cell(row=r, column=idx_col).value
                    if value:
                        max_len = max(max_len, min(len(str(value)), 50))
                ws.column_dimensions[get_column_letter(idx_col)].width = max(10, max_len + 2)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"✅ XLSX создан: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Конвертация ВБК PDF → XLSX")
    parser.add_argument("pdf_path", help="Путь к PDF ВБК")
    parser.add_argument("output_path", nargs="?", help="Путь к XLSX (опционально)")
    args = parser.parse_args()

    pdf_path = Path(args.pdf_path)
    if not pdf_path.exists():
        raise SystemExit(f"❌ PDF не найден: {pdf_path}")

    if args.output_path:
        output_path = Path(args.output_path)
    else:
        output_dir = Path("output") / pdf_path.stem
        output_path = output_dir / f"{pdf_path.stem}.xlsx"

    print("📂 Входной PDF:", pdf_path)
    print("📂 Выходной XLSX:", output_path)

    _, grouped = extract_tables_with_sections(pdf_path)
    build_workbook(grouped, output_path)
    print("🎉 Конвертация завершена успешно!")


if __name__ == "__main__":
    main()

