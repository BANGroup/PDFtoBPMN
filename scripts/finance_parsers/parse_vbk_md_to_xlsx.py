#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Парсер ВБК из MD → XLSX

Преобразует MD файл с таблицами ВБК (созданный run_document.py) 
в структурированный XLSX файл с человекопонятными названиями.

Usage:
    python3 scripts/utils/parse_vbk_md_to_xlsx.py output/VBK.../VBK..._OCR.md output/VBK.../VBK.xlsx
"""

import sys
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Импортировать справочник полей
sys.path.insert(0, str(Path(__file__).parent))
from vbk_field_names import get_field_name, get_table_name


def parse_md_tables(md_path: Path) -> dict:
    """
    Извлекает таблицы из MD файла.
    
    Returns:
        dict с ключами:
            - 'metadata': метаданные из frontmatter
            - 'tables': список словарей {'section': str, 'data': list[list]}
    """
    with open(md_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    result = {
        'metadata': {},
        'tables': []
    }
    
    # Извлечь метаданные из frontmatter (опционально)
    frontmatter_match = re.search(r'^---\n(.*?)\n---', content, re.DOTALL)
    if frontmatter_match:
        # Простое извлечение - можно улучшить с YAML парсером
        result['metadata']['raw'] = frontmatter_match.group(1)
    
    # Найти все таблицы в MD (формат: | col1 | col2 |)
    # Разделить по страницам
    pages = re.split(r'<!-- Страница \d+ -->', content)
    
    current_section = "Раздел I. Учетная информация"
    
    for page_num, page_content in enumerate(pages, 1):
        # Найти заголовки разделов
        section_match = re.search(r'^[#]+\s+(Раздел .+?)(?:\s*\{#.*?\})?$', page_content, re.MULTILINE)
        if section_match:
            current_section = section_match.group(1).strip()
        
        # Найти все таблицы на странице
        # Паттерн: строка с | ... | ... | (заголовок)
        #          строка с | --- | --- | (разделитель)
        #          строки с данными
        table_pattern = r'\|[^\n]+\|\n\|[\s\-:]+\|\n(?:\|[^\n]+\|\n?)+'
        
        for table_match in re.finditer(table_pattern, page_content):
            table_text = table_match.group(0)
            
            # Разобрать таблицу
            lines = [line.strip() for line in table_text.strip().split('\n') if line.strip()]
            
            if len(lines) < 3:  # Минимум: заголовок + разделитель + 1 строка данных
                continue
            
            # Заголовок
            header = [cell.strip() for cell in lines[0].split('|')[1:-1]]
            
            # Данные (пропуск разделителя - строка 1)
            data_rows = []
            for line in lines[2:]:
                cells = [cell.strip() for cell in line.split('|')[1:-1]]
                if len(cells) == len(header):
                    data_rows.append(cells)
            
            if data_rows:
                result['tables'].append({
                    'section': current_section,
                    'page': page_num,
                    'headers': header,
                    'data': data_rows
                })
    
    return result


def identify_vbk_table_type(section: str, headers: list) -> str:
    """
    Определяет тип таблицы ВБК по разделу и заголовкам.
    
    Returns:
        Код таблицы (Table1, Table2, Table6...) или 'Unknown'
    """
    section_lower = section.lower()
    
    # Раздел I
    if 'нерезидент' in section_lower or 'реквизиты нерезидента' in section_lower:
        return 'Table1'
    
    if 'контракт' in section_lower and 'общие сведения' in section_lower:
        return 'Table2'
    
    if 'постановка на учет' in section_lower or 'регистрационный номер банка' in section_lower:
        return 'Table3'
    
    # Раздел II - Платежи
    if 'платеж' in section_lower or 'раздел ii' in section_lower:
        # Проверить заголовки
        headers_str = ' '.join(headers).lower()
        if any(x in headers_str for x in ['код вида операции', 'сумма', 'валюта']):
            return 'Table6'
    
    # Раздел III - Подтверждающие документы
    if 'подтверждающ' in section_lower or 'раздел iii' in section_lower:
        return 'Table4'
    
    # Раздел VII - Зачет
    if 'зачет' in section_lower or 'встречн' in section_lower:
        return 'Table7'
    
    # Раздел X - Корректировки
    if 'корректиров' in section_lower:
        return 'Table10'
    
    return 'Unknown'


def write_vbk_xlsx(parsed_data: dict, output_path: Path):
    """
    Создает XLSX файл из распарсенных данных ВБК.
    """
    wb = Workbook()
    
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Титульный лист
    ws_title = wb.create_sheet("ВБК")
    ws_title.merge_cells('A1:D1')
    title_cell = ws_title['A1']
    title_cell.value = "ВЕДОМОСТЬ БАНКОВСКОГО КОНТРОЛЯ"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws_title.append([])
    ws_title.append(["Извлечено из PDF", ""])
    ws_title.append(["Таблиц найдено:", len(parsed_data['tables'])])
    
    ws_title.column_dimensions['A'].width = 25
    ws_title.column_dimensions['B'].width = 60
    
    # Группировать таблицы по типу
    tables_by_type = {}
    for table in parsed_data['tables']:
        table_type = identify_vbk_table_type(table['section'], table['headers'])
        
        if table_type not in tables_by_type:
            tables_by_type[table_type] = []
        
        tables_by_type[table_type].append(table)
    
    # Создать листы для каждого типа таблицы
    for table_type in sorted(tables_by_type.keys()):
        tables = tables_by_type[table_type]
        
        # Название листа
        if table_type != 'Unknown':
            readable_name = get_table_name(table_type)
            sheet_name = readable_name[:31]
        else:
            sheet_name = f"Прочие таблицы ({len(tables)})"
        
        ws = wb.create_sheet(sheet_name)
        
        # Заголовок таблицы
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        title_cell = ws.cell(1, 1)
        title_cell.value = sheet_name
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Объединить все таблицы этого типа
        all_headers = set()
        all_data = []
        
        for table in tables:
            all_headers.update(table['headers'])
            
            # Добавить данные с выравниванием по заголовкам
            for row in table['data']:
                row_dict = dict(zip(table['headers'], row))
                all_data.append(row_dict)
        
        headers_list = sorted(all_headers)
        
        # Записать заголовки (строка 3)
        ws.append([])
        ws.append(headers_list)
        
        # Форматирование заголовков
        for cell in ws[3]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_thin
        
        # Записать данные
        for row_dict in all_data:
            row = [row_dict.get(header, '') for header in headers_list]
            ws.append(row)
        
        # Границы
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=len(headers_list)):
            for cell in row:
                cell.border = border_thin
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Автоширина
        for idx, col in enumerate(headers_list, start=1):
            max_length = len(col)
            for row_idx in range(4, ws.max_row + 1):
                cell_value = str(ws.cell(row_idx, idx).value or '')
                max_length = max(max_length, min(len(cell_value), 50))
            
            ws.column_dimensions[chr(64 + idx)].width = max_length + 2
    
    wb.save(output_path)
    return tables_by_type


def main():
    """
    Основная функция.
    """
    if len(sys.argv) < 2:
        print("Использование: python3 scripts/utils/parse_vbk_md_to_xlsx.py <input_md> [output_xlsx]")
        print()
        print("Примеры:")
        print("  python3 scripts/utils/parse_vbk_md_to_xlsx.py output/VBK.../VBK..._OCR.md")
        print("  python3 scripts/utils/parse_vbk_md_to_xlsx.py output/VBK.../VBK..._OCR.md output/VBK.xlsx")
        sys.exit(1)
    
    input_path = Path(sys.argv[1])
    
    if not input_path.exists():
        print(f"❌ Файл не найден: {input_path}")
        sys.exit(1)
    
    # Определить output путь
    if len(sys.argv) >= 3:
        output_path = Path(sys.argv[2])
    else:
        output_path = input_path.parent / (input_path.stem.replace('_OCR', '') + '.xlsx')
    
    print(f"📂 Входной MD:  {input_path}")
    print(f"📂 Выходной XLSX: {output_path}")
    print()
    
    # Парсинг MD
    print("📄 Парсинг MD файла...")
    try:
        parsed_data = parse_md_tables(input_path)
    except Exception as e:
        print(f"❌ Ошибка парсинга MD: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    
    print(f"   ✅ Таблиц найдено: {len(parsed_data['tables'])}")
    print()
    
    # Создание XLSX
    print("📊 Создание XLSX...")
    try:
        tables_by_type = write_vbk_xlsx(parsed_data, output_path)
        print(f"   ✅ XLSX создан: {output_path}")
    except Exception as e:
        print(f"❌ Ошибка создания XLSX: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    
    print()
    print("📋 Типы таблиц:")
    for table_type in sorted(tables_by_type.keys()):
        count = len(tables_by_type[table_type])
        name = get_table_name(table_type) if table_type != 'Unknown' else "Прочие"
        print(f"   • {table_type:10} ({count:3} таблиц) - {name}")
    
    print()
    print("🎉 Конвертация завершена успешно!")


if __name__ == "__main__":
    main()




