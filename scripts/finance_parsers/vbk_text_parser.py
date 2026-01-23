#!/usr/bin/env python3
"""
Улучшенный парсер для VBK документа.
Использует текстовое содержимое PDF вместо find_tables() для точного извлечения.
"""

import fitz
import pandas as pd
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import numbers


class VBKTextParser:
    """Парсер на основе текстового содержимого"""
    
    def __init__(self, pdf_path: str, section: str = "II"):
        self.pdf_path = pdf_path
        self.section = section  # "II" или "III"
        self.doc = None
        self.all_records = []
        
        if section == "II":
            self.num_columns = 19
            self.financial_columns = [
                "Сумма операции (платеж) - сумма",
                "Сумма операции (контракт) - сумма"
            ]
            self.headers = [
                "№ п/п",
                "Дата операции",
                "Направление (признак) платежа",
                "Признак совершения операции третьим лицом",
                "Код вида операции",
                "Код валюты корр.счета",
                "Сумма операции (платеж) - код валюты",
                "Сумма операции (платеж) - сумма",
                "Сумма операции (контракт) - код валюты",
                "Сумма операции (контракт) - сумма",
                "Ожидаемый срок репатриации",
                "Код страны банка получателя/отправителя",
                "Банк-нерезидент - код страны",
                "Банк-нерезидент - наименование",
                "Банк-нерезидент - код банка",
                "Банк-нерезидент - номер счета",
                "Признак изменения записи",
                "Признак представления документов",
                "Примечание"
            ]
        else:  # section == "III"
            self.num_columns = 15
            self.financial_columns = [
                "Сумма по документам (документ) - сумма",
                "Сумма по документам (контракт) - сумма"
            ]
            self.headers = [
                "№ п/п",
                "Подтверждающий документ - номер",
                "Подтверждающий документ - дата",
                "Код вида подтверждающего документа",
                "Признак исполнения обязательств третьим лицом",
                "Сумма по документам (документ) - код валюты",
                "Сумма по документам (документ) - сумма",
                "Сумма по документам (контракт) - код валюты",
                "Сумма по документам (контракт) - сумма",
                "Признак поставки",
                "Срок исполнения",
                "Признак изменения записи",
                "Код страны грузоотправителя/грузополучателя",
                "Дополнительная информация",
                "Примечание"
            ]
    
    def parse(self):
        """Извлечь данные из PDF"""
        self.doc = fitz.open(self.pdf_path)
        
        print(f"📄 Обработка файла: {self.pdf_path}")
        print(f"📊 Раздел: {self.section}")
        print(f"📊 Всего страниц: {len(self.doc)}")
        print()
        
        # Находим стартовую страницу
        start_page = self._find_section_start()
        if start_page is None:
            print(f"❌ Не найден 'Раздел {self.section}' в документе")
            return pd.DataFrame()
        
        print(f"✅ Раздел {self.section} найден на странице {start_page + 1}")
        print()
        
        # Извлекаем данные со всех страниц
        page_num = start_page
        while page_num < len(self.doc):
            should_continue = self._extract_from_page(page_num)
            if not should_continue:
                print(f"  ⚠️ Конец Раздела {self.section}")
                break
            page_num += 1
        
        print(f"\n📍 Раздел {self.section}: страницы {start_page + 1} - {page_num}")
        
        self.doc.close()
        
        # Преобразуем в DataFrame
        if not self.all_records:
            print("❌ Не извлечено ни одной записи")
            return pd.DataFrame()
        
        df = pd.DataFrame(self.all_records, columns=self.headers)
        df = self._clean_data(df)
        
        print(f"\n📊 Извлечено записей: {len(df)}")
        
        return df
    
    def _find_section_start(self) -> int:
        """Найти страницу с началом раздела"""
        marker = f"Раздел {self.section}"
        for page_num in range(len(self.doc)):
            page = self.doc[page_num]
            text = page.get_text()
            if marker in text:
                # Для Раздела III начинаем со следующей страницы
                if self.section == "III":
                    return page_num + 1 if page_num + 1 < len(self.doc) else page_num
                return page_num
        return None
    
    def _extract_from_page(self, page_num: int) -> bool:
        """
        Извлечь данные с одной страницы.
        Возвращает True если нужно продолжать, False если конец раздела.
        """
        page = self.doc[page_num]
        text = page.get_text()
        lines = text.split('\n')
        
        # Фильтруем пустые строки
        lines = [l.strip() for l in lines if l.strip()]
        
        # Ищем записи (начинаются с числа - номера п/п)
        records_found = 0
        i = 0
        
        while i < len(lines):
            line = lines[i]
            
            # Проверяем, является ли это началом новой записи (число)
            if re.match(r'^\d+$', line):
                row_num = int(line)
                
                # Пытаемся извлечь следующие значения для этой записи
                record = self._extract_record(lines, i)
                
                if record and len(record) >= self.num_columns - 5:  # Допускаем неполные записи
                    # Дополняем до нужного количества колонок
                    while len(record) < self.num_columns:
                        record.append(None)
                    
                    self.all_records.append(record[:self.num_columns])
                    records_found += 1
                
                # Переходим к следующему потенциальному номеру
                i += 1
            else:
                i += 1
        
        if records_found > 0:
            print(f"  Страница {page_num + 1}: извлечено {records_found} записей")
            return True
        else:
            # Если на странице нет записей, возможно конец раздела
            return False
    
    def _extract_record(self, lines: list, start_idx: int) -> list:
        """
        Извлечь одну запись начиная с start_idx.
        Возвращает список значений для всех колонок.
        """
        record = []
        idx = start_idx
        
        # Колонка 0: номер п/п
        if idx < len(lines) and re.match(r'^\d+$', lines[idx]):
            record.append(int(lines[idx]))
            idx += 1
        else:
            return []
        
        # Собираем следующие значения
        # Для Раздела II: дата, направление, признак, код операции, код валюты, суммы...
        # Для Раздела III: номер документа, дата, код вида, признак, суммы...
        
        values_to_collect = self.num_columns - 1  # Минус уже собранный номер п/п
        collected = 0
        multi_line_text = ""
        
        while idx < len(lines) and collected < values_to_collect:
            line = lines[idx]
            
            # Если встретили следующий номер записи - останавливаемся
            if re.match(r'^\d+$', line) and collected >= values_to_collect - 3:
                # Это может быть следующая запись
                break
            
            # Паттерны для определения типа значения
            is_date = bool(re.match(r'^\d{2}\.\d{2}\.\d{4}$', line))
            is_number = bool(re.match(r'^\d+$', line))
            is_decimal = bool(re.match(r'^\d+[,.]?\d*$', line.replace(',', '')))
            is_code = bool(re.match(r'^[A-Z0-9_-]+$', line))
            
            # Если это явное значение - добавляем
            if is_date or is_number or is_decimal or is_code:
                if multi_line_text:
                    record.append(multi_line_text.strip())
                    multi_line_text = ""
                    collected += 1
                
                record.append(line)
                collected += 1
            else:
                # Это может быть многострочный текст
                if multi_line_text:
                    multi_line_text += " " + line
                else:
                    multi_line_text = line
            
            idx += 1
            
            # Защита от бесконечного цикла
            if idx - start_idx > 100:
                break
        
        # Добавляем последний накопленный текст если есть
        if multi_line_text and collected < values_to_collect:
            record.append(multi_line_text.strip())
            collected += 1
        
        return record
    
    def _clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Очистить и форматировать данные"""
        # Преобразуем финансовые колонки в float
        for col in self.financial_columns:
            if col in df.columns:
                df[col] = df[col].apply(self._parse_number)
        
        # Очищаем пустые значения
        df = df.replace(['None', 'nan', ''], None)
        
        return df
    
    def _parse_number(self, value):
        """Преобразовать строку в число"""
        if pd.isna(value) or value is None or value == '':
            return None
        
        value_str = str(value).strip()
        if not value_str or value_str in ['None', 'nan', '']:
            return None
        
        # Убираем пробелы и заменяем запятую на точку
        value_str = value_str.replace(' ', '').replace(',', '')
        
        try:
            return float(value_str)
        except ValueError:
            return None
    
    def save_to_excel(self, df: pd.DataFrame, output_path: str):
        """Сохранить в Excel с правильными форматами"""
        df.to_excel(output_path, index=False, engine='openpyxl')
        
        # Применяем форматирование
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Находим индексы финансовых колонок
        header_row = list(df.columns)
        financial_col_indices = []
        
        for col_name in self.financial_columns:
            if col_name in header_row:
                col_idx = header_row.index(col_name) + 1
                financial_col_indices.append(col_idx)
        
        # Применяем финансовый формат с 2 знаками
        for col_idx in financial_col_indices:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is not None:
                    cell.number_format = numbers.FORMAT_NUMBER_00
        
        wb.save(output_path)
        print(f"\n💾 Сохранено в: {output_path}")
        print(f"   Применен финансовый формат для {len(financial_col_indices)} колонок")

