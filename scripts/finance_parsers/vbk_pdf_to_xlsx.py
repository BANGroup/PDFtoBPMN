#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ВБК PDF → XLSX (прямой парсер)

Использует пайплайн для извлечения таблиц из PDF ВБК
и создает структурированный XLSX файл.

Usage:
    python3 scripts/utils/vbk_pdf_to_xlsx.py input/VBK.pdf output/VBK.xlsx
"""

import sys
from pathlib import Path

# Добавить путь к пайплайну
sys.path.insert(0, str(Path(__file__).parent.parent))

from pdf_to_context.document_pipeline import DocumentToContextPipeline
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Импортировать справочник полей
sys.path.insert(0, str(Path(__file__).parent))
from vbk_field_names import get_table_name


def extract_tables_from_pdf(pdf_path: Path) -> list:
    """
    Извлекает таблицы из PDF используя пайплайн.
    """
    from pdf_to_context.extractors.pdf_extractor import PDFExtractor
    
    extractor = PDFExtractor()
    
    # Извлечь блоки из PDF
    blocks = extractor.extract_document(str(pdf_path))
    
    # Отфильтровать только таблицы
    tables = []
    
    for block in blocks:
        if block.get('type') == 'table':
            # Преобразовать content в list[list]
            table_data = block.get('content', [])
            
            tables.append({
                'page': block.get('page', 0),
                'data': table_data,
                'context_before': '',
                'context_after': ''
            })
    
    return tables


def identify_vbk_section(context_text: str) -> str:
    """
    Определяет раздел ВБК по контексту.
    """
    context_lower = context_text.lower()
    
    if 'раздел i' in context_lower or 'учетная информация' in context_lower:
        if 'нерезидент' in context_lower:
            return 'Раздел I. Нерезидент'
        elif 'контракт' in context_lower:
            return 'Раздел I. Контракт'
        elif 'постановка на учет' in context_lower:
            return 'Раздел I. Постановка на учет'
        else:
            return 'Раздел I. Учетная информация'
    
    if 'раздел ii' in context_lower or 'платеж' in context_lower:
        return 'Раздел II. Сведения о платежах'
    
    if 'раздел iii' in context_lower or 'подтверждающ' in context_lower:
        return 'Раздел III. Подтверждающие документы'
    
    if 'раздел iv' in context_lower:
        return 'Раздел IV'
    
    if 'раздел v' in context_lower or 'итоговые данные' in context_lower:
        return 'Раздел V. Итоговые данные'
    
    if 'зачет' in context_lower or 'встречн' in context_lower:
        return 'Раздел VII. Зачет встречных требований'
    
    return 'Прочие таблицы'


def write_vbk_xlsx(tables: list, output_path: Path):
    """
    Создает XLSX файл из таблиц ВБК.
    """
    wb = Workbook()
    
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    title_font = Font(bold=True, size=14)
    title_alignment = Alignment(horizontal="center", vertical="center")
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Титульный лист
    ws_title = wb.create_sheet("ВБК")
    ws_title.merge_cells('A1:D1')
    title_cell = ws_title['A1']
    title_cell.value = "ВЕДОМОСТЬ БАНКОВСКОГО КОНТРОЛЯ"
    title_cell.font = title_font
    title_cell.alignment = title_alignment
    
    ws_title.append([])
    ws_title.append(["Извлечено из PDF", ""])
    ws_title.append(["Таблиц найдено:", len(tables)])
    
    ws_title.column_dimensions['A'].width = 25
    ws_title.column_dimensions['B'].width = 60
    
    # Группировать таблицы по разделам
    tables_by_section = {}
    for table in tables:
        context = table['context_before'] + ' ' + table['context_after']
        section = identify_vbk_section(context)
        
        if section not in tables_by_section:
            tables_by_section[section] = []
        
        tables_by_section[section].append(table)
    
    # Создать листы для каждого раздела
    for section in sorted(tables_by_section.keys()):
        tables_in_section = tables_by_section[section]
        
        # Название листа (обрезать до 31 символа)
        sheet_name = section[:31]
        ws = wb.create_sheet(sheet_name)
        
        # Заголовок раздела
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        title_cell = ws.cell(1, 1)
        title_cell.value = section
        title_cell.font = title_font
        title_cell.alignment = title_alignment
        
        ws.append([])  # Пустая строка
        
        # Записать все таблицы раздела
        for table_idx, table in enumerate(tables_in_section):
            table_data = table['data']
            
            if not table_data or len(table_data) < 2:
                continue
            
            # Заголовок таблицы (если несколько таблиц в разделе)
            if len(tables_in_section) > 1:
                ws.append([f"Таблица {table_idx + 1} (стр. {table['page']})"])
                ws.cell(ws.max_row, 1).font = Font(bold=True)
            
            # Первая строка - заголовки
            headers = table_data[0]
            ws.append(headers)
            
            # Форматирование заголовков
            for cell in ws[ws.max_row]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_thin
            
            # Данные
            for row in table_data[1:]:
                ws.append(row)
            
            # Границы для данных
            data_start_row = ws.max_row - len(table_data) + 2
            data_end_row = ws.max_row
            
            for row in ws.iter_rows(min_row=data_start_row, max_row=data_end_row, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.border = border_thin
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Автоширина колонок
            for idx in range(1, len(headers) + 1):
                max_length = len(str(headers[idx-1]))
                for row_idx in range(data_start_row, data_end_row + 1):
                    cell_value = str(ws.cell(row_idx, idx).value or '')
                    max_length = max(max_length, min(len(cell_value), 50))
                
                ws.column_dimensions[get_column_letter(idx)].width = max_length + 2
            
            # Пустая строка между таблицами
            ws.append([])
    
    wb.save(output_path)
    return tables_by_section


def main():
    """
    Основная функция.
    """
    if len(sys.argv) < 2:
        print("Использование: python3 scripts/utils/vbk_pdf_to_xlsx.py <input_pdf> [output_xlsx]")
        print()
        print("Примеры:")
        print("  python3 scripts/utils/vbk_pdf_to_xlsx.py input/VBK.pdf")
        print("  python3 scripts/utils/vbk_pdf_to_xlsx.py input/VBK.pdf output/VBK.xlsx")
        sys.exit(1)
    
    input_path = Path(sys.argv[1])
    
    if not input_path.exists():
        print(f"❌ Файл не найден: {input_path}")
        sys.exit(1)
    
    # Определить output путь
    if len(sys.argv) >= 3:
        output_path = Path(sys.argv[2])
    else:
        output_dir = input_path.parent.parent / "output" / input_path.stem
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / (input_path.stem + ".xlsx")
    
    print(f"📂 Входной PDF:  {input_path}")
    print(f"📂 Выходной XLSX: {output_path}")
    print()
    
    # Извлечение таблиц
    print("📄 Извлечение таблиц из PDF...")
    try:
        tables = extract_tables_from_pdf(input_path)
    except Exception as e:
        print(f"❌ Ошибка извлечения: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    
    print(f"   ✅ Таблиц найдено: {len(tables)}")
    print()
    
    # Создание XLSX
    print("📊 Создание XLSX...")
    try:
        tables_by_section = write_vbk_xlsx(tables, output_path)
        print(f"   ✅ XLSX создан: {output_path}")
    except Exception as e:
        print(f"❌ Ошибка создания XLSX: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    
    print()
    print("📋 Разделы:")
    for section in sorted(tables_by_section.keys()):
        count = len(tables_by_section[section])
        print(f"   • {section:40} ({count} таблиц)")
    
    print()
    print("🎉 Конвертация завершена успешно!")


if __name__ == "__main__":
    main()

