#!/usr/bin/env python3
"""
Парсер для извлечения таблицы из Раздела III документа VBK.
Извлекает таблицу со всех страниц и сохраняет в Excel с правильными форматами.
"""

import fitz
import pandas as pd
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import numbers


class VBKSection3Parser:
    """Парсер для Раздела III (Сведения о подтверждающих документах)"""
    
    def __init__(self, pdf_path: str):
        self.pdf_path = pdf_path
        self.doc = None
        self.all_data = []
        
        # Определяем колонки с финансовыми данными
        self.financial_columns = [
            "Сумма по документам (документ) - сумма",
            "Сумма по документам (контракт) - сумма"
        ]
        
    def parse(self):
        """Извлечь таблицу из всех страниц Раздела III"""
        self.doc = fitz.open(self.pdf_path)
        
        print(f"📄 Обработка файла: {self.pdf_path}")
        print(f"📊 Всего страниц: {len(self.doc)}")
        print()
        
        # Находим страницу с началом Раздела III
        start_page = self._find_section3_start()
        if start_page is None:
            print("❌ Не найден 'Раздел III' в документе")
            return []
        
        print(f"✅ Раздел III найден на странице {start_page + 1}")
        print()
        
        # Извлекаем таблицы со всех страниц начиная с Раздела III
        # Раздел III продолжается до конца документа (или до следующего раздела)
        page_num = start_page
        while page_num < len(self.doc):
            # Пытаемся извлечь таблицу
            should_continue = self._extract_table_from_page(page_num)
            
            if not should_continue:
                print(f"  ⚠️ Структура таблицы изменилась - конец Раздела III")
                break
            
            page_num += 1
        
        print(f"\n📍 Раздел III: страницы {start_page + 1} - {page_num}")
        
        self.doc.close()
        
        # Собираем все данные в один DataFrame
        if not self.all_data:
            print("❌ Не извлечено ни одной строки данных")
            return pd.DataFrame()
        
        df = pd.concat(self.all_data, ignore_index=True)
        
        # Очищаем данные
        df = self._clean_data(df)
        
        print(f"\n📊 Извлечено строк: {len(df)}")
        print(f"📊 Колонок: {len(df.columns)}")
        
        return df
    
    def _find_section3_start(self) -> int:
        """Найти страницу с началом Раздела III"""
        for page_num in range(len(self.doc)):
            page = self.doc[page_num]
            text = page.get_text()
            if 'Раздел III' in text:
                # Начинаем со следующей страницы, где уже будет полная структура таблицы
                return page_num + 1 if page_num + 1 < len(self.doc) else page_num
        return None
    
    def _extract_table_from_page(self, page_num: int) -> bool:
        """
        Извлечь таблицу с одной страницы.
        
        Возвращает:
            True - если таблица имеет структуру Раздела III (14-15 колонок), продолжаем
            False - если структура изменилась, останавливаемся
        """
        page = self.doc[page_num]
        
        tab_finder = page.find_tables()
        tables = tab_finder.tables
        
        if not tables:
            print(f"  Страница {page_num + 1}: таблица не найдена")
            return True  # Продолжаем искать
        
        # Берем только первую таблицу на странице
        table = tables[0]
        df = table.to_pandas()
        
        # Проверяем структуру таблицы
        num_cols = len(df.columns)
        
        # Раздел III: 14-15 колонок
        if num_cols < 13 or num_cols > 16:
            # Структура изменилась - это уже не Раздел III
            return False
        
        # Переименовываем колонки в числовые индексы для унификации
        df.columns = range(len(df.columns))
        
        # Фильтруем строки (убираем заголовки и пустые)
        # ИСПРАВЛЕНИЕ СМЕЩЕНИЯ КОЛОНОК И ФИЛЬТРАЦИЯ ЗАГОЛОВКОВ
        cleaned_rows = []
        for idx, row in df.iterrows():
            first_col = str(row.iloc[0]).strip()
            
            if not first_col or first_col in ['None', '', 'nan', '№ п/п']:
                continue
            
            # Проверяем, что это не строка с номерами колонок
            # УТОЧНЕНИЕ: если во второй или третьей колонке есть дата или номер документа - это НЕ заголовок
            if first_col in ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', 
                              '11', '12', '13', '14', '15'] and idx < 5:
                # Проверяем вторую и третью колонки на признаки данных
                second_col = str(row.iloc[1] if len(row) > 1 else '').strip()
                third_col = str(row.iloc[2] if len(row) > 2 else '').strip()
                
                has_date = bool(re.match(r'\d{2}\.\d{2}\.\d{4}', third_col))
                has_doc_number = bool(re.match(r'[A-Z]\d+', second_col))
                
                if not (has_date or has_doc_number):
                    # Нет признаков данных - это заголовок, пропускаем
                    continue
            
            # ИСПРАВЛЕНИЕ СМЕЩЕНИЯ: если Col 1 = None/nan, а Col 2 похоже на данные - удаляем Col 1
            if len(row) > 2:
                col1_val = str(row.iloc[1]).strip()
                col2_val = str(row.iloc[2]).strip()
                
                # Проверяем паттерны данных
                is_col2_data = bool(re.match(r'[A-Z0-9]', col2_val))
                
                if col1_val in ['None', 'nan', ''] and is_col2_data:
                    # Смещение обнаружено - удаляем Col 1
                    row = pd.concat([row.iloc[:1], row.iloc[2:]], ignore_index=True)
            
            # Обрезаем до 15 колонок (максимальная структура)
            if len(row) > 15:
                row = row[:15]
            
            cleaned_rows.append(row)
        
        if cleaned_rows:
            page_df = pd.DataFrame(cleaned_rows)
            # Убеждаемся что у нас ровно 15 колонок
            if len(page_df.columns) < 15:
                for i in range(len(page_df.columns), 15):
                    page_df[i] = None
            elif len(page_df.columns) > 15:
                page_df = page_df.iloc[:, :15]
            
            self.all_data.append(page_df)
            print(f"  Страница {page_num + 1}: извлечено {len(cleaned_rows)} строк")
        
        return True  # Продолжаем извлекать
    
    def _clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Очистить и форматировать данные"""
        # Устанавливаем правильные названия колонок
        headers = [
            "№ п/п",
            "Подтверждающий документ - номер",
            "Подтверждающий документ - дата",
            "Код вида подтверждающего документа",
            "Признак исполнения обязательств третьим лицом",
            "Сумма по документам (документ) - код валюты",
            "Сумма по документам (документ) - сумма",
            "Сумма по документам (контракт) - код валюты",
            "Сумма по документам (контракт) - сумма",
            "Признак поставки",
            "Срок исполнения",
            "Признак изменения записи",
            "Код страны грузоотправителя/грузополучателя",
            "Дополнительная информация",
            "Примечание"
        ]
        
        # DataFrame уже должен иметь 15 колонок (числовые индексы 0-14)
        if len(df.columns) == 15:
            df.columns = headers
        else:
            print(f"⚠️ WARNING: Ожидалось 15 колонок, получено {len(df.columns)}")
            if len(df.columns) < 15:
                for i in range(len(df.columns), 15):
                    df[i] = None
            elif len(df.columns) > 15:
                df = df.iloc[:, :15]
            df.columns = headers
        
        # Преобразуем финансовые колонки в float
        for col in self.financial_columns:
            if col in df.columns:
                df[col] = df[col].apply(self._parse_number)
        
        # Очищаем пустые значения
        df = df.replace(['None', 'nan', ''], None)
        
        return df
    
    def _parse_number(self, value):
        """Преобразовать строку в число"""
        if pd.isna(value) or value is None or value == '':
            return None
        
        value_str = str(value).strip()
        if not value_str or value_str in ['None', 'nan', '']:
            return None
        
        # Убираем пробелы и заменяем запятую на точку
        value_str = value_str.replace(' ', '').replace(',', '')
        
        try:
            return float(value_str)
        except ValueError:
            return None
    
    def save_to_excel(self, df: pd.DataFrame, output_path: str):
        """Сохранить в Excel с правильными форматами"""
        # Сохраняем в Excel
        df.to_excel(output_path, index=False, engine='openpyxl')
        
        # Применяем форматирование
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Находим индексы финансовых колонок
        header_row = list(df.columns)
        financial_col_indices = []
        
        for col_name in self.financial_columns:
            if col_name in header_row:
                col_idx = header_row.index(col_name) + 1  # +1 т.к. Excel 1-indexed
                financial_col_indices.append(col_idx)
        
        # Применяем финансовый формат с 2 знаками
        for col_idx in financial_col_indices:
            for row in range(2, ws.max_row + 1):  # Пропускаем заголовок
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is not None:
                    cell.number_format = numbers.FORMAT_NUMBER_00  # Финансовый формат с 2 знаками
        
        wb.save(output_path)
        print(f"\n💾 Сохранено в: {output_path}")
        print(f"   Применен финансовый формат для {len(financial_col_indices)} колонок")


if __name__ == '__main__':
    pdf_path = "input/VBK16040002_1971_0019_9_1_2216_0008_20251111_большая.pdf"
    output_path = "output/finance/VBK_Раздел_III.xlsx"
    
    # Создаем выходную папку
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    
    # Парсим
    parser = VBKSection3Parser(pdf_path)
    df = parser.parse()
    
    if not df.empty:
        parser.save_to_excel(df, output_path)
        
        # Статистика
        print("\n" + "="*80)
        print("📊 СТАТИСТИКА:")
        print(f"   Всего строк: {len(df)}")
        print(f"   Всего колонок: {len(df.columns)}")
        print(f"   Финансовые колонки: {parser.financial_columns}")

