"""
Загрузка каталога документов из xlsx файла БНД
Содержит метаданные: процесс, тип, дата регистрации, статус
"""

import re
from pathlib import Path
from typing import Dict, Optional
from dataclasses import dataclass
from datetime import datetime


@dataclass
class CatalogEntry:
    """Запись из каталога БНД"""
    unid: str
    doc_num: str                           # Код документа (РД-М1.014-16)
    doc_type: Optional[str] = None         # Тип (РД - Руководство по деятельности)
    process: Optional[str] = None          # Процесс (М1 - Анализ и оценка)
    reg_date: Optional[datetime] = None    # Дата регистрации
    status: Optional[str] = None           # Статус (Действующий/Архивный)
    created: Optional[datetime] = None     # Дата создания
    executor: Optional[str] = None         # Исполнитель


def load_catalog(xlsx_path: Path) -> Dict[str, CatalogEntry]:
    """
    Загрузить каталог из xlsx файла
    
    Returns:
        Dict[doc_code.upper() -> CatalogEntry]
    """
    try:
        import openpyxl
    except ImportError:
        return {}
    
    catalog = {}
    
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb.active
        
        # Получаем заголовки
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        header_idx = {h: i for i, h in enumerate(headers) if h}
        
        # Читаем строки
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                unid = row[header_idx.get('UNID', 0)]
                doc_num = row[header_idx.get('DocNum', 7)]
                
                if not doc_num:
                    continue
                
                entry = CatalogEntry(
                    unid=unid or '',
                    doc_num=doc_num,
                    doc_type=row[header_idx.get('Type', 11)] if 'Type' in header_idx else None,
                    process=row[header_idx.get('Process', 10)] if 'Process' in header_idx else None,
                    reg_date=row[header_idx.get('DocRegDate', 8)] if 'DocRegDate' in header_idx else None,
                    status=row[header_idx.get('DocStatus', 9)] if 'DocStatus' in header_idx else None,
                    created=row[header_idx.get('Created', 1)] if 'Created' in header_idx else None,
                    executor=row[header_idx.get('DocExecutorSNM', 4)] if 'DocExecutorSNM' in header_idx else None,
                )
                
                # Нормализуем код для поиска
                code_normalized = normalize_code(doc_num)
                catalog[code_normalized] = entry
                
            except (IndexError, KeyError):
                continue
        
        wb.close()
        
    except Exception as e:
        print(f"⚠️ Ошибка загрузки каталога: {e}")
    
    return catalog


def normalize_code(code: str) -> str:
    """Нормализация кода документа для сопоставления"""
    if not code:
        return ""
    
    # Приводим к верхнему регистру
    code = code.upper()
    
    # Убираем пробелы
    code = code.replace(' ', '')
    
    # Унифицируем разделители
    code = code.replace('/', '-').replace('.', '-')
    
    return code


def find_in_catalog(catalog: Dict[str, CatalogEntry], doc_code: str) -> Optional[CatalogEntry]:
    """
    Найти документ в каталоге по коду
    
    Пробует разные варианты нормализации
    """
    if not doc_code:
        return None
    
    # Прямой поиск
    normalized = normalize_code(doc_code)
    if normalized in catalog:
        return catalog[normalized]
    
    # Поиск по частичному совпадению (без версии)
    base_code = re.sub(r'-\d+$', '', doc_code)
    normalized_base = normalize_code(base_code)
    
    for key, entry in catalog.items():
        if key.startswith(normalized_base):
            return entry
    
    return None


def extract_process_name(process_str: str) -> tuple:
    """
    Извлечь код и название процесса
    
    "М1 - Анализ и оценка" -> ("М1", "Анализ и оценка")
    """
    if not process_str:
        return None, None
    
    match = re.match(r'^([МБВ]\d+(?:\.\d+)?)\s*-\s*(.+)$', process_str)
    if match:
        return match.group(1), match.group(2).strip()
    
    return None, process_str


def extract_type_name(type_str: str) -> tuple:
    """
    Извлечь код и название типа
    
    "РД - Руководство по деятельности" -> ("РД", "Руководство по деятельности")
    """
    if not type_str:
        return None, None
    
    match = re.match(r'^([А-ЯA-Z]{2,5})\s*-\s*(.+)$', type_str)
    if match:
        return match.group(1), match.group(2).strip()
    
    return None, type_str


if __name__ == "__main__":
    import sys
    
    xlsx_path = Path("/home/budnik_an/Obligations/input2/BND/ИС БНД выгрузка 2025.12.26.xlsx")
    
    print(f"📊 Загрузка каталога: {xlsx_path.name}")
    print("=" * 60)
    
    catalog = load_catalog(xlsx_path)
    print(f"✅ Загружено записей: {len(catalog)}")
    
    # Тест поиска
    test_codes = ['РД-М1.014-16', 'ДП-Б1.004-06', 'СТ-166-01', 'РК01-2017-07']
    
    print("\n🔍 Тест поиска:")
    for code in test_codes:
        entry = find_in_catalog(catalog, code)
        if entry:
            proc_code, proc_name = extract_process_name(entry.process)
            type_code, type_name = extract_type_name(entry.doc_type)
            print(f"\n📄 {code}")
            print(f"   Процесс: {proc_code} - {proc_name}")
            print(f"   Тип: {type_code} - {type_name}")
            print(f"   Дата: {entry.reg_date}")
            print(f"   Статус: {entry.status}")
        else:
            print(f"\n❌ {code} - не найден")
