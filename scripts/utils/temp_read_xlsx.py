
import sys
import openpyxl
from pathlib import Path

def read_excel(file_path):
    print(f"--- Reading {file_path} ---")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            print(f"Sheet: {sheet_name}")
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                # Filter out completely empty rows
                if any(row):
                    cleaned_row = [str(cell) if cell is not None else "" for cell in row]
                    print(" | ".join(cleaned_row))
            print("-" * 20)
    except Exception as e:
        print(f"Error reading {file_path}: {e}")

if __name__ == "__main__":
    files = [
        "input2/Salym/Приложение 1 закупки и договоры -04.12.25.xlsx",
        "input2/Salym/Приложение 3 Хронология по договорам 08.12.2025.xlsx"
    ]
    for f in files:
        read_excel(f)

