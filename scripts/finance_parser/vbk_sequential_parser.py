#!/usr/bin/env python3
"""
Последовательный парсер для VBK документа.
Группирует PDF-строки по номерам п/п и склеивает многострочные значения.
"""

import fitz
import pandas as pd
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import numbers
from typing import List, Dict, Any


class VBKSequentialParser:
    """Парсер с группировкой по номерам п/п"""
    
    def __init__(self, pdf_path: str, section: str = "II"):
        self.pdf_path = pdf_path
        self.section = section
        self.doc = None
        self.records = {}  # {row_num: {col_idx: [values]}}
        
        if section == "II":
            self.num_columns = 19
            self.financial_col_indices = [7, 9]  # Суммы (1-indexed: 8, 10)
            self.headers = [
                "№ п/п",
                "Дата операции",
                "Направление (признак) платежа",
                "Признак совершения операции третьим лицом",
                "Код вида операции",
                "Код валюты корр.счета",
                "Сумма операции (платеж) - код валюты",
                "Сумма операции (платеж) - сумма",
                "Сумма операции (контракт) - код валюты",
                "Сумма операции (контракт) - сумма",
                "Ожидаемый срок репатриации",
                "Код страны банка получателя/отправителя",
                "Банк-нерезидент - код страны",
                "Банк-нерезидент - наименование",
                "Банк-нерезидент - код банка",
                "Банк-нерезидент - номер счета",
                "Признак изменения записи",
                "Признак представления документов",
                "Примечание"
            ]
        else:  # section == "III"
            self.num_columns = 15
            self.financial_col_indices = [6, 8]
            self.headers = [
                "№ п/п",
                "Подтверждающий документ - номер",
                "Подтверждающий документ - дата",
                "Код вида подтверждающего документа",
                "Признак исполнения обязательств третьим лицом",
                "Сумма по документам (документ) - код валюты",
                "Сумма по документам (документ) - сумма",
                "Сумма по документам (контракт) - код валюты",
                "Сумма по документам (контракт) - сумма",
                "Признак поставки",
                "Срок исполнения",
                "Признак изменения записи",
                "Код страны грузоотправителя/грузополучателя",
                "Дополнительная информация",
                "Примечание"
            ]
    
    def parse(self):
        """Извлечь данные из PDF"""
        self.doc = fitz.open(self.pdf_path)
        
        print(f"📄 Обработка файла: {self.pdf_path}")
        print(f"📊 Раздел: {self.section}")
        print(f"📊 Всего страниц: {len(self.doc)}")
        print()
        
        # Находим стартовую страницу
        start_page = self._find_section_start()
        if start_page is None:
            print(f"❌ Не найден 'Раздел {self.section}' в документе")
            return pd.DataFrame()
        
        print(f"✅ Раздел {self.section} найден на странице {start_page + 1}")
        print()
        
        # Извлекаем весь текст со всех страниц раздела
        all_lines = []
        page_num = start_page
        
        while page_num < len(self.doc):
            page = self.doc[page_num]
            text = page.get_text()
            
            # Проверяем конец раздела
            if page_num > start_page:
                expected_cols = 19 if self.section == "II" else 15
                other_cols = 15 if self.section == "II" else 19
                
                # Если структура таблицы изменилась - конец раздела
                tab_finder = page.find_tables()
                if tab_finder.tables:
                    num_cols = len(tab_finder.tables[0].to_pandas().columns)
                    if abs(num_cols - other_cols) < abs(num_cols - expected_cols):
                        print(f"  ⚠️ Структура таблицы изменилась на странице {page_num + 1}")
                        break
            
            lines = text.split('\n')
            all_lines.extend([(page_num + 1, line.strip()) for line in lines if line.strip()])
            page_num += 1
        
        print(f"📍 Раздел {self.section}: страницы {start_page + 1} - {page_num}")
        print(f"📝 Всего строк текста: {len(all_lines)}")
        print()
        
        self.doc.close()
        
        # Группируем строки по номерам п/п
        self._group_by_row_numbers(all_lines)
        
        # Преобразуем в DataFrame
        df = self._build_dataframe()
        
        print(f"\n📊 Извлечено записей: {len(df)}")
        
        return df
    
    def _find_section_start(self) -> int:
        """Найти страницу с началом раздела"""
        marker = f"Раздел {self.section}"
        for page_num in range(len(self.doc)):
            page = self.doc[page_num]
            text = page.get_text()
            if marker in text:
                if self.section == "III":
                    return page_num + 1 if page_num + 1 < len(self.doc) else page_num
                return page_num
        return None
    
    def _group_by_row_numbers(self, all_lines: List[tuple]):
        """
        Группировать строки PDF по номерам п/п.
        
        Алгоритм:
        1. Ищем номер п/п (число от 1 до 9999)
        2. Все строки до следующего номера = продолжение текущей записи
        3. Склеиваем значения колонок
        """
        current_row_num = None
        current_values = []
        
        print("🔄 Группировка строк по номерам п/п...")
        
        for page_num, line in all_lines:
            # Проверяем, является ли это номером п/п (просто число)
            if re.match(r'^\d{1,4}$', line):
                # Это потенциальный номер п/п
                row_num = int(line)
                
                # Если это новый номер - сохраняем предыдущую запись
                if current_row_num is not None and row_num != current_row_num:
                    self._save_record(current_row_num, current_values)
                    current_values = []
                
                current_row_num = row_num
                current_values.append(line)  # Добавляем сам номер
            else:
                # Это продолжение текущей записи
                if current_row_num is not None:
                    current_values.append(line)
        
        # Сохраняем последнюю запись
        if current_row_num is not None:
            self._save_record(current_row_num, current_values)
        
        print(f"  ✅ Обработано записей: {len(self.records)}")
    
    def _save_record(self, row_num: int, values: List[str]):
        """
        Сохранить запись.
        
        values - список всех строк PDF, относящихся к этой записи.
        Нужно распределить их по колонкам.
        """
        if row_num not in self.records:
            self.records[row_num] = {}
        
        # Распределяем значения по колонкам
        # Первое значение - номер п/п (колонка 0)
        self.records[row_num][0] = [str(row_num)]
        
        # Остальные значения распределяем по паттернам
        col_idx = 1
        i = 1  # Пропускаем первое значение (номер)
        
        while i < len(values) and col_idx < self.num_columns:
            val = values[i]
            
            # Определяем тип значения по паттернам
            is_date = bool(re.match(r'^\d{2}\.\d{2}\.\d{4}$', val))
            is_integer = bool(re.match(r'^\d{1,6}$', val))
            is_decimal = bool(re.match(r'^\d{1,10}[,.]?\d{0,2}$', val.replace(',', '')))
            is_code = bool(re.match(r'^[A-Z0-9_-]{2,10}$', val))
            
            # Распределение по колонкам для Раздела II
            if self.section == "II":
                if col_idx == 1 and is_date:  # Дата операции
                    self.records[row_num][col_idx] = [val]
                    col_idx += 1
                    i += 1
                elif col_idx in [2, 3, 4, 5] and is_integer:  # Коды и признаки
                    self.records[row_num][col_idx] = [val]
                    col_idx += 1
                    i += 1
                elif col_idx in [6, 8] and is_integer:  # Коды валют
                    self.records[row_num][col_idx] = [val]
                    col_idx += 1
                    i += 1
                elif col_idx in [7, 9] and is_decimal:  # Суммы
                    self.records[row_num][col_idx] = [val]
                    col_idx += 1
                    i += 1
                elif col_idx >= 10:  # Текстовые поля
                    # Собираем все оставшиеся строки как текст
                    if col_idx not in self.records[row_num]:
                        self.records[row_num][col_idx] = []
                    self.records[row_num][col_idx].append(val)
                    
                    # Если это короткое значение (код) - переходим к следующей колонке
                    if len(val) < 10 and (is_integer or is_code):
                        col_idx += 1
                    
                    i += 1
                else:
                    # Пропускаем неподходящее значение
                    i += 1
            else:  # Раздел III
                # Аналогичная логика для Раздела III
                if col_idx == 1 and is_code:  # Номер документа
                    self.records[row_num][col_idx] = [val]
                    col_idx += 1
                    i += 1
                elif col_idx == 2 and is_date:  # Дата документа
                    self.records[row_num][col_idx] = [val]
                    col_idx += 1
                    i += 1
                elif col_idx in [3, 4, 5, 7] and is_code:  # Коды
                    self.records[row_num][col_idx] = [val]
                    col_idx += 1
                    i += 1
                elif col_idx in [6, 8] and is_decimal:  # Суммы
                    self.records[row_num][col_idx] = [val]
                    col_idx += 1
                    i += 1
                elif col_idx >= 9:  # Текстовые поля
                    if col_idx not in self.records[row_num]:
                        self.records[row_num][col_idx] = []
                    self.records[row_num][col_idx].append(val)
                    
                    if len(val) < 10 and (is_integer or is_code):
                        col_idx += 1
                    
                    i += 1
                else:
                    i += 1
    
    def _build_dataframe(self) -> pd.DataFrame:
        """Построить DataFrame из сгруппированных записей"""
        rows = []
        
        for row_num in sorted(self.records.keys()):
            row_data = []
            
            for col_idx in range(self.num_columns):
                if col_idx in self.records[row_num]:
                    # Склеиваем значения через пробел
                    values = self.records[row_num][col_idx]
                    merged = ' '.join(values).strip()
                    row_data.append(merged)
                else:
                    row_data.append(None)
            
            rows.append(row_data)
        
        df = pd.DataFrame(rows, columns=self.headers)
        
        # Преобразуем типы данных
        df = self._convert_types(df)
        
        return df
    
    def _convert_types(self, df: pd.DataFrame) -> pd.DataFrame:
        """Преобразовать типы данных"""
        # Преобразуем финансовые колонки
        for col_idx in self.financial_col_indices:
            col_name = self.headers[col_idx]
            df[col_name] = df[col_name].apply(self._parse_number)
        
        # Очищаем пустые значения
        df = df.replace(['None', 'nan', ''], None)
        
        return df
    
    def _parse_number(self, value):
        """Преобразовать строку в число"""
        if pd.isna(value) or value is None or value == '':
            return None
        
        value_str = str(value).strip()
        if not value_str or value_str in ['None', 'nan', '']:
            return None
        
        # Убираем пробелы и заменяем запятую на точку
        value_str = value_str.replace(' ', '').replace(',', '')
        
        try:
            return float(value_str)
        except ValueError:
            return None
    
    def save_to_excel(self, df: pd.DataFrame, output_path: str):
        """Сохранить в Excel с правильными форматами"""
        df.to_excel(output_path, index=False, engine='openpyxl')
        
        # Применяем форматирование
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Применяем финансовый формат
        for col_idx in self.financial_col_indices:
            excel_col = col_idx + 1  # Excel 1-indexed
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=excel_col)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = numbers.FORMAT_NUMBER_00
        
        wb.save(output_path)
        print(f"\n💾 Сохранено в: {output_path}")
        print(f"   Применен финансовый формат для {len(self.financial_col_indices)} колонок")

