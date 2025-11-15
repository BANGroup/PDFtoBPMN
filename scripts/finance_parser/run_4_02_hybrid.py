#!/usr/bin/env python3
"""
Гибридный парсер для Выпуска 4-02:
- MD: ФИО, адреса, документы
- PDF: количества (структура отличается от 4-01)
"""

import sys
import re
import fitz
from pathlib import Path
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from scripts.finance_parser.md_parser_4_02 import MDParser

def extract_quantities_from_pdf(pdf_path: Path) -> dict:
    """Извлекает количества из PDF (специфичная структура 4-02)"""
    doc = fitz.open(pdf_path)
    
    quantities = {}
    
    for page_num in range(len(doc)):
        page = doc[page_num]
        text = page.get_text()
        lines = text.split('\n')
        
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            
            # Ищем "Количество в штуках"
            if "Количество в штуках" in line:
                # Следующая строка должна быть числом
                if i + 1 < len(lines):
                    qty_line = lines[i + 1].strip()
                    qty_match = re.match(r'^(\d{1,7})$', qty_line)
                    
                    if qty_match:
                        qty = int(qty_match.group(1))
                        
                        # Ищем код владельца: сначала ВПЕРЕД (i+2 до i+50)
                        code = None
                        for j in range(i + 2, min(len(lines), i + 50)):
                            code_match = re.search(r'(01_\d{11}|02_\d{11}|03_\d{11})', lines[j])
                            if code_match:
                                code = code_match.group(1)
                                break
                        
                        # Если не нашли вперед - ищем НАЗАД (i-20 до i-1)
                        if not code:
                            for j in range(i - 1, max(0, i - 30), -1):
                                code_match = re.search(r'(01_\d{11}|02_\d{11}|03_\d{11})', lines[j])
                                if code_match:
                                    code = code_match.group(1)
                                    break
                        
                        # Сохраняем только если нашли код и его еще нет
                        if code and code not in quantities:
                            quantities[code] = qty
            
            i += 1
    
    doc.close()
    
    return quantities


def main():
    """Главная функция"""
    print("="*80)
    print("🔍 ГИБРИДНАЯ ОБРАБОТКА: Выпуск 4-02 на 16.06.2020")
    print("📋 MD: ФИО, адреса, документы")
    print("📋 PDF: количества")
    print("="*80)
    print()
    
    # Пути
    md_path = Path("output/Выпуск_4-02_на_16.06.2020/Выпуск_4-02_на_16.06.2020_OCR.md")
    pdf_path = Path("input/Finance/Выпуск 4-02 на 16.06.2020.pdf")
    output_path = Path("output/finance/Выпуск_4-02_на_16.06.2020.xlsx")
    
    # 1. Извлекаем владельцев из MD
    print("📖 Парсинг MD (ФИО, адреса, документы)...")
    parser = MDParser()
    records = parser.parse_md_file(str(md_path))
    print(f"   ✅ Извлечено записей: {len(records)}")
    
    # 2. Извлекаем количества из PDF
    print("📄 Извлечение количеств из PDF...")
    quantities = extract_quantities_from_pdf(pdf_path)
    print(f"   ✅ Извлечено количеств: {len(quantities)}")
    print(f"   📊 Сумма: {sum(quantities.values()):,}")
    print()
    
    # 3. Обогащаем записи количествами
    print("🔗 Обогащение записей количествами...")
    enriched = 0
    for rec in records:
        if rec.owner_code in quantities:
            rec.quantity = quantities[rec.owner_code]
            enriched += 1
    
    print(f"   ✅ Обогащено: {enriched}/{len(records)}")
    print()
    
    # Статистика
    total_qty = sum(r.quantity for r in records if r.quantity)
    filled_qty = sum(1 for r in records if r.quantity)
    filled_fio = sum(1 for r in records if r.full_name)
    filled_addr = sum(1 for r in records if r.address)
    filled_doc = sum(1 for r in records if r.document_number)
    
    print("🔍 Валидация данных...")
    print(f"   Записей: {len(records)}")
    print(f"   Облигаций: {total_qty:,}")
    print(f"   Ожидается: 9,179,259")
    print(f"   Разница: {total_qty - 9179259:+,} ({100*(total_qty - 9179259)/9179259:+.2f}%)")
    print()
    print("   Заполненность:")
    print(f"   • Количество:  {100*filled_qty/len(records):.1f}%")
    print(f"   • ФИО:         {100*filled_fio/len(records):.1f}%")
    print(f"   • Адрес:       {100*filled_addr/len(records):.1f}%")
    print(f"   • Документ:    {100*filled_doc/len(records):.1f}%")
    print()
    
    # Экспорт
    print("📦 Создание Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Реестр владельцев"
    
    # Заголовки
    headers = ['Адрес регистрации', 'Количество в штуках', 'Код владельца', 
               'ФИО', 'Номер документа', 'Номер счета', 'Страница']
    ws.append(headers)
    
    # Данные
    for rec in records:
        ws.append([
            rec.address or '',
            rec.quantity or 0,
            rec.owner_code or '',
            rec.full_name or '',
            rec.document_number or '',
            rec.account_number or '',
            rec.page_number or 0
        ])
    
    # Форматирование
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '0'
    
    wb.save(output_path)
    
    print(f"✅ ГОТОВО: {output_path}")
    print()
    print("="*80)


if __name__ == '__main__':
    main()

