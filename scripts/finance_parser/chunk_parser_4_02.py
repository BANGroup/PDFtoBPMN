#!/usr/bin/env python3
"""
Chunk-based парсер для Выпуск 4-02
Разбивает PDF на чанки по маркеру "Код, присвоенный номинальным держателем"
"""

import re
import fitz
from pathlib import Path
from typing import List, Optional
from dataclasses import dataclass

@dataclass
class OwnerRecord:
    owner_code: str
    full_name: Optional[str] = None
    address: Optional[str] = None
    document_number: Optional[str] = None
    quantity: Optional[int] = None
    account_number: Optional[str] = None
    page_number: Optional[int] = None


class ChunkParser:
    """Парсер разбивающий PDF на чанки по владельцам"""
    
    def parse(self, pdf_path: Path) -> List[OwnerRecord]:
        """Парсит PDF разбивая на чанки"""
        doc = fitz.open(pdf_path)
        
        # Собираем весь текст
        full_text = ""
        page_offsets = [0]  # Смещения начала каждой страницы
        
        for page_num in range(len(doc)):
            page_text = doc[page_num].get_text()
            full_text += page_text
            page_offsets.append(len(full_text))
        
        doc.close()
        
        # Разбиваем на чанки по маркеру
        marker = "Код, присвоенный номинальным держателем,"
        
        chunks = []
        start = 0
        
        while True:
            # Ищем следующий маркер
            next_pos = full_text.find(marker, start + len(marker))
            
            if next_pos == -1:
                # Последний чанк
                if start < len(full_text):
                    chunks.append(full_text[start:])
                break
            else:
                chunks.append(full_text[start:next_pos])
                start = next_pos
        
        print(f"📦 Разбито на {len(chunks)} чанков")
        print()
        
        # Парсим каждый чанк
        records = []
        for i, chunk in enumerate(chunks):
            if not chunk.strip():
                continue
            
            record = self._parse_chunk(chunk, i + 1)
            if record:
                records.append(record)
        
        return records
    
    def _parse_chunk(self, chunk: str, chunk_num: int) -> Optional[OwnerRecord]:
        """Парсит один чанк владельца"""
        
        # 1. Извлекаем код владельца
        code_match = re.search(r'предоставляющим данные\s+(\d{2}_\d{11})', chunk)
        if not code_match:
            return None
        
        owner_code = code_match.group(1)
        
        # 2. Извлекаем количество (в КОНЦЕ ТЕКУЩЕГО чанка, ПЕРЕД следующим кодом)
        # Количество для текущего владельца находится ПОСЛЕ его кода, но ДО следующего кода
        quantity = None
        
        # Ищем "Количество в штуках" в текущем чанке ПОСЛЕ кода владельца
        # ВАЖНО: может быть ДВА числа - берем число ПЕРЕД прописью!
        chunk_after_code = chunk[code_match.end():]
        
        # Сначала пытаемся найти число прямо перед "(прописью)"
        qty_match_with_words = re.search(r'Количество в штуках[\s\S]{0,200}?(\d{1,7})[\s\n]+[а-яё\s]+[\s\n]*\(прописью\)', chunk_after_code)
        
        if qty_match_with_words:
            # Нашли с прописью - это точное количество
            quantity = int(qty_match_with_words.group(1))
        else:
            # Нет прописи - берем первое число
            qty_match = re.search(r'Количество в штуках[\s\S]{0,50}?(\d{1,7})', chunk_after_code)
            if qty_match:
                quantity = int(qty_match.group(1))
        
        # 3. Извлекаем ФИО/название
        fio = None
        fio_match = re.search(r'Почтовое наименование\s+(.+?)(?=\s*Почтовый адрес)', chunk, re.DOTALL)
        if fio_match:
            fio_raw = fio_match.group(1).strip()
            # Очищаем
            fio = ' '.join(fio_raw.split())
            if len(fio) > 200:
                fio = fio[:197] + '...'
        
        # 4. Извлекаем адрес
        # Адрес может начинаться с кода страны (2 буквы), потом ЦИФРА (индекс) или ТЕКСТ (для иностранных)
        address = None
        addr_match = re.search(r'Почтовый адрес\s+([A-Z]{2}\s+.+?)(?=\s*Код страны местонахождения)', chunk, re.DOTALL)
        if addr_match:
            addr_raw = addr_match.group(1).strip()
            address = ' '.join(addr_raw.split())
        
        # 5. Извлекаем номер документа
        # После "Номер и/или серия документа" идет "Дата документа", 
        # потом еще текст, а затем сам номер
        # Номер может быть:
        #   - Только цифры (10-20): российские юрлица (ОГРН, ИНН)
        #   - Буквы+цифры: иностранные компании (НЕ390071, HE390071)
        document_number = None
        
        # Ищем секцию с документом
        # У юрлиц: "Номер и/или серия документа" (одна строка)
        # У физлиц: "Номер и/или серия\nдокумента" (с переносом!)
        doc_marker = re.search(r'Номер и/или серия[\s\n]+документа', chunk)
        if doc_marker:
            # Берем текст после маркера (до 500 символов)
            doc_section = chunk[doc_marker.end():doc_marker.end()+500]
            
            # Сначала пытаемся найти буквенно-цифровой (иностранные компании)
            # Формат: 2+ буквы + пробел + цифры (HE 205891)
            doc_match = re.search(r'\b([A-ZА-Я]{2,}[\s]?\d{5,15})\b', doc_section)
            if doc_match:
                raw = doc_match.group(1)
                document_number = raw.replace(' ', '')  # Убираем пробелы
            else:
                # Затем пытаемся найти цифры (с возможными пробелами для паспортов)
                # Паттерн: начинается с цифры, может содержать ТОЛЬКО пробелы (не \n!), заканчивается цифрой
                # Примеры: "4512 703546", "45 02 057362", "1027700070419", "8040031", "970"
                # ВАЖНО: используем [ ] вместо \s чтобы НЕ захватывать переносы строк
                doc_match = re.search(r'\b(\d[\d ]{4,25}\d|\d{3,20})\b', doc_section)
                if doc_match:
                    raw_number = doc_match.group(1)
                    # Очищаем от пробелов
                    clean_number = raw_number.replace(' ', '')
                    # Проверяем что это действительно номер (только цифры, 3+ символов)
                    if clean_number.isdigit() and len(clean_number) >= 3:
                        document_number = clean_number
        
        # 6. Номер счета
        account_number = None
        acc_match = re.search(r'Номер счета\s+(\S+)', chunk)
        if acc_match:
            account_number = acc_match.group(1)
        
        return OwnerRecord(
            owner_code=owner_code,
            full_name=fio,
            address=address,
            document_number=document_number,
            quantity=quantity,
            account_number=account_number,
            page_number=chunk_num
        )
    
    def save_to_excel(self, records: List[OwnerRecord], output_path: Path):
        """Сохраняет записи в Excel"""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Реестр владельцев"
        
        headers = ['Адрес регистрации', 'Количество в штуках', 'Код владельца', 
                   'ФИО', 'Номер документа', 'Номер счета', 'Страница']
        ws.append(headers)
        
        for rec in records:
            ws.append([
                rec.address or '',
                rec.quantity or 0,
                rec.owner_code or '',
                rec.full_name or '',
                rec.document_number or '',
                rec.account_number or '',
                rec.page_number or 0
            ])
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.number_format = '0'
        
        wb.save(output_path)


if __name__ == '__main__':
    from openpyxl import Workbook
    
    print("="*80)
    print("🔍 CHUNK-BASED ПАРСЕР: Выпуск 4-02 на 16.06.2020")
    print("="*80)
    print()
    
    pdf_path = Path("input/Finance/Выпуск 4-02 на 16.06.2020.pdf")
    output_path = Path("output/finance/Выпуск_4-02_на_16.06.2020.xlsx")
    
    parser = ChunkParser()
    records = parser.parse(pdf_path)
    
    print(f"✅ Извлечено записей: {len(records)}")
    print()
    
    # Статистика
    total_qty = sum(r.quantity for r in records if r.quantity)
    filled_qty = sum(1 for r in records if r.quantity)
    filled_fio = sum(1 for r in records if r.full_name)
    filled_addr = sum(1 for r in records if r.address)
    filled_doc = sum(1 for r in records if r.document_number)
    
    print("🔍 Валидация данных...")
    print(f"   Записей:       {len(records)}")
    print(f"   Облигаций:     {total_qty:,}")
    print(f"   Ожидается:     9,179,259")
    print(f"   Разница:       {total_qty - 9179259:+,} ({100*(total_qty - 9179259)/9179259:+.2f}%)")
    print()
    print("   Заполненность:")
    print(f"   • Количество:  {100*filled_qty/len(records):.1f}%")
    print(f"   • ФИО:         {100*filled_fio/len(records):.1f}%")
    print(f"   • Адрес:       {100*filled_addr/len(records):.1f}%")
    print(f"   • Документ:    {100*filled_doc/len(records):.1f}%")
    print()
    
    # Экспорт
    print("📦 Создание Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Реестр владельцев"
    
    headers = ['Адрес регистрации', 'Количество в штуках', 'Код владельца', 
               'ФИО', 'Номер документа', 'Номер счета', 'Страница']
    ws.append(headers)
    
    for rec in records:
        ws.append([
            rec.address or '',
            rec.quantity or 0,
            rec.owner_code or '',
            rec.full_name or '',
            rec.document_number or '',
            rec.account_number or '',
            rec.page_number or 0
        ])
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '0'
    
    wb.save(output_path)
    
    print(f"✅ ГОТОВО: {output_path}")
    print()
    print("="*80)

