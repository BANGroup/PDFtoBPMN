#!/usr/bin/env python3
"""
Обработка Выпуск 4-02 на 16.06.2020
Использует рабочий парсер от 4-01 на 16.06.2020 (формат по дате!)
"""

import sys
from pathlib import Path
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from scripts.finance_parser.md_parser import MDParser

def main():
    print("="*80)
    print("🔍 ОБРАБОТКА: Выпуск 4-02 на 16.06.2020")
    print("📋 Парсер: рабочий от 4-01 на 16.06.2020 (формат по дате!)")
    print("="*80)
    print()
    
    md_path = Path("output/Выпуск_4-02_на_16.06.2020/Выпуск_4-02_на_16.06.2020_OCR.md")
    output_path = Path("output/finance/Выпуск_4-02_на_16.06.2020.xlsx")
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    print("📖 Парсинг MD файла...")
    parser = MDParser()
    records = parser.parse_md_file(str(md_path))
    print(f"   ✅ Извлечено записей: {len(records)}")
    print()
    
    # Статистика
    total_qty = sum(r.quantity for r in records if r.quantity)
    filled_qty = sum(1 for r in records if r.quantity)
    filled_fio = sum(1 for r in records if r.full_name)
    filled_addr = sum(1 for r in records if r.address)
    filled_doc = sum(1 for r in records if r.document_number)
    
    print("🔍 Валидация данных...")
    print(f"   Записей: {len(records)}")
    print(f"   Облигаций: {total_qty:,}")
    print(f"   Ожидается: 9,179,259")
    print(f"   Разница: {total_qty - 9179259:+,} ({100*(total_qty - 9179259)/9179259:+.2f}%)")
    print()
    print("   Заполненность:")
    print(f"   • Количество:  {100*filled_qty/len(records):.1f}%")
    print(f"   • ФИО:         {100*filled_fio/len(records):.1f}%")
    print(f"   • Адрес:       {100*filled_addr/len(records):.1f}%")
    print(f"   • Документ:    {100*filled_doc/len(records):.1f}%")
    print()
    
    # Экспорт в Excel
    print("📦 Создание Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Реестр владельцев"
    
    headers = ['Адрес регистрации', 'Количество в штуках', 'Код владельца', 
               'ФИО', 'Номер документа', 'Номер счета', 'Страница']
    ws.append(headers)
    
    for rec in records:
        ws.append([
            rec.address or '',
            rec.quantity or 0,
            rec.owner_code or '',
            rec.full_name or '',
            rec.document_number or '',
            rec.account_number or '',
            rec.page_number or 0
        ])
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '0'
    
    wb.save(output_path)
    
    print(f"✅ ГОТОВО: {output_path}")
    print()
    print("="*80)

if __name__ == '__main__':
    main()


