#!/usr/bin/env python3
"""
Простой парсер с присваиванием по порядку: владелец[i] → количество[i]
"""

import re
import fitz
from pathlib import Path
from typing import List, Optional
from dataclasses import dataclass
from openpyxl import Workbook

@dataclass
class OwnerRecord:
    page_number: int
    owner_code: str
    full_name: Optional[str] = None
    address: Optional[str] = None
    document_number: Optional[str] = None
    account_number: Optional[str] = None
    quantity: Optional[int] = None

class SimpleOrderParser:
    """Парсер с простым присваиванием по порядку"""
    
    def parse(self, pdf_path: Path) -> List[OwnerRecord]:
        """Парсит PDF извлекая владельцев и количества отдельно"""
        print("🔍 ПАРСИНГ С ПРИСВАИВАНИЕМ ПО ПОРЯДКУ")
        print("="*80)
        print()
        
        doc = fitz.open(pdf_path)
        
        # ШАГ 1: Извлекаем владельцев
        print("📋 ШАГ 1: Извлечение владельцев...")
        owners = self._extract_owners(doc)
        print(f"   ✅ Найдено: {len(owners)}")
        
        # ШАГ 2: Извлекаем количества
        print("📊 ШАГ 2: Извлечение количеств...")
        quantities = self._extract_quantities(doc)
        print(f"   ✅ Найдено: {len(quantities)} (сумма: {sum(quantities):,})")
        
        doc.close()
        
        # ШАГ 3: Присваиваем по порядку
        print("🔗 ШАГ 3: Присваивание по порядку (1:1)...")
        for i, owner in enumerate(owners):
            if i < len(quantities):
                owner.quantity = quantities[i]
        
        filled = sum(1 for o in owners if o.quantity)
        print(f"   ✅ Заполнено: {filled}/{len(owners)}")
        print()
        
        return owners
    
    def _extract_owners(self, doc) -> List[OwnerRecord]:
        """Извлекает владельцев в порядке появления"""
        owners = []
        
        for page_num in range(len(doc)):
            page = doc[page_num]
            text = page.get_text()
            lines = text.split('\n')
            
            # Находим коды владельцев
            owner_matches = list(re.finditer(r'(01_\d{10})\(NADC\)', text))
            
            for match_idx, match in enumerate(owner_matches):
                owner_code = match.group(1)
                record = OwnerRecord(
                    page_number=page_num + 1,
                    owner_code=owner_code
                )
                
                # Чанк после кода
                start = match.start()
                next_match = owner_matches[match_idx + 1] if match_idx + 1 < len(owner_matches) else None
                chunk_forward = text[start:next_match.start()] if next_match else text[start:]
                
                # Чанк до кода
                prev_match = owner_matches[match_idx - 1] if match_idx > 0 else None
                chunk_back = text[prev_match.start():start] if prev_match else text[:start]
                
                # Извлекаем поля
                record.full_name = self._extract_fio(chunk_forward, lines, match.start(), text) or \
                                   self._extract_fio(chunk_back, lines, match.start(), text)
                record.address = self._extract_address(chunk_forward) or self._extract_address(chunk_back)
                record.document_number = self._extract_document(chunk_forward) or self._extract_document(chunk_back)
                record.account_number = self._extract_account(chunk_forward)
                
                owners.append(record)
        
        return owners
    
    def _extract_quantities(self, doc) -> List[int]:
        """Извлекает количества в порядке появления"""
        quantities = []
        
        for page_num in range(len(doc)):
            page = doc[page_num]
            text = page.get_text()
            lines = text.split('\n')
            
            for i, line in enumerate(lines):
                if '4-01-36484-R' in line:
                    if i + 2 < len(lines):
                        qty_line = lines[i + 2].strip()
                        # ✅ ВСЕГДА берем число на строке i+2
                        qty_match = re.search(r'^\d{1,7}$', qty_line)
                        if qty_match:
                            qty = int(qty_match.group())
                            if 1 <= qty < 3000000:
                                quantities.append(qty)
        
        return quantities
    
    def _extract_fio(self, chunk: str, lines: List[str], code_pos: int, full_text: str) -> Optional[str]:
        """Извлекает ФИО"""
        chunk_lines = chunk.split('\n')
        
        # Паттерн 1: "Полное наименование/"
        for i, line in enumerate(chunk_lines):
            if 'Полное наименование' in line:
                for offset in [1, 2]:
                    if i - offset >= 0:
                        fio_line = chunk_lines[i - offset].strip()
                        if fio_line and len(fio_line) > 5:
                            if '(LEI)' not in fio_line and 'Ф.И.О.' not in fio_line and \
                               fio_line not in ['Код', 'Счет', 'Номер'] and not fio_line.startswith('Номер '):
                                return fio_line
                break
        
        # Паттерн 2: "Ф.И.О. руководителя"
        for i, line in enumerate(chunk_lines):
            if 'Ф.И.О. руководителя' in line:
                for offset in [1, 2]:
                    if i - offset >= 0:
                        fio_line = chunk_lines[i - offset].strip()
                        if fio_line and len(fio_line) > 10:
                            if not any(kw in fio_line for kw in ['ОГРН', 'Код ОКПО', 'Код ОКВЭД', 
                                                                  'Регистр', 'Адрес', 'Контакт']):
                                return fio_line
                break
        
        # Паттерн 3: "Адрес для направления"
        for i, line in enumerate(chunk_lines):
            if 'Адрес для направления' in line:
                for offset in [2, 3, 4]:
                    if i - offset >= 0:
                        fio_line = chunk_lines[i - offset].strip()
                        if fio_line and len(fio_line) > 10:
                            if not any(kw in fio_line for kw in ['ОГРН', 'Код', 'Ф.И.О.', 'Регистр', 'Контакт']):
                                return fio_line
                break
        
        return None
    
    def _extract_address(self, chunk: str) -> Optional[str]:
        """Извлекает адрес"""
        # Паттерн 1: "RU РОССИЯ [индекс] [адрес]"
        pattern = r'RU\s+РОССИЯ\s+\d{6}\s+([^\n]+?)(?:\s+Адрес|\n)'
        match = re.search(pattern, chunk, re.IGNORECASE)
        if match:
            addr = match.group(1).strip()
            addr = re.sub(r'\s{2,}', ' ', addr)
            if len(addr) > 10:
                return addr
        
        # Паттерн 2: "Адрес для направления корреспонденции"
        lines = chunk.split('\n')
        for i, line in enumerate(lines):
            if 'Адрес для направления' in line:
                if i + 2 < len(lines):
                    addr_line = lines[i + 2].strip()
                    if addr_line and len(addr_line) > 10:
                        if any(c.isdigit() for c in addr_line):
                            return addr_line
                break
        
        return None
    
    def _extract_document(self, chunk: str) -> Optional[str]:
        """Извлекает номер документа"""
        # Паспорт
        pattern1 = r'(?:паспорт|Паспорт).*?(\d{2}\s*\d{2}\s*\d{6})'
        match = re.search(pattern1, chunk, re.IGNORECASE)
        if match:
            doc_num = match.group(1).replace(' ', '')
            if len(doc_num) == 10:
                return f"{doc_num[:2]} {doc_num[2:4]} {doc_num[4:]}"
        
        # ОГРН
        pattern2 = r'ОГРН\s+(\d{13})'
        match = re.search(pattern2, chunk, re.IGNORECASE)
        if match:
            return match.group(1)
        
        # ИНН
        pattern3 = r'ИНН\s+(\d{10,12})'
        match = re.search(pattern3, chunk, re.IGNORECASE)
        if match:
            return match.group(1)
        
        return None
    
    def _extract_account(self, chunk: str) -> Optional[str]:
        """Извлекает номер счета"""
        pattern = r'Номер\s+([A-Z0-9_/]+)\s+Тип счета'
        match = re.search(pattern, chunk, re.IGNORECASE)
        if match:
            return match.group(1).strip()
        return None


def main():
    """Главная функция"""
    pdf_path = Path("input/Finance/Выпуск 4-01 на 29.07.19.pdf")
    output_path = Path("output/finance/Выпуск_4-01_29_07_19_PERFECT.xlsx")
    
    parser = SimpleOrderParser()
    records = parser.parse(pdf_path)
    
    print("="*80)
    print("📊 СТАТИСТИКА:")
    print("="*80)
    
    total_qty = sum(r.quantity for r in records if r.quantity)
    filled_fio = sum(1 for r in records if r.full_name)
    filled_addr = sum(1 for r in records if r.address)
    filled_doc = sum(1 for r in records if r.document_number)
    
    print(f"   Записей:       {len(records)}")
    print(f"   Облигаций:     {total_qty:,}")
    print(f"   Ожидается:     4,121,600")
    print(f"   Разница:       {total_qty - 4121600:+,} ({100*(total_qty - 4121600)/4121600:+.3f}%)")
    print()
    print("📋 ЗАПОЛНЕННОСТЬ:")
    print(f"   Количество:    {sum(1 for r in records if r.quantity)}/{len(records)}")
    print(f"   ФИО:           {filled_fio}/{len(records)}")
    print(f"   Адрес:         {filled_addr}/{len(records)}")
    print(f"   Документ:      {filled_doc}/{len(records)}")
    print()
    
    # Экспорт в Excel
    print("📦 Создание Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Реестр владельцев"
    
    # Заголовки
    headers = ['Адрес регистрации', 'Количество в штуках', 'Код владельца', 
               'ФИО', 'Номер документа', 'Номер счета', 'Страница']
    ws.append(headers)
    
    # Данные
    for rec in records:
        ws.append([
            rec.address or '',
            rec.quantity or 0,
            rec.owner_code or '',
            rec.full_name or '',
            rec.document_number or '',
            rec.account_number or '',
            rec.page_number or 0
        ])
    
    # Форматирование
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '0'
    
    wb.save(output_path)
    
    print(f"✅ ГОТОВО: {output_path}")


if __name__ == '__main__':
    main()

